import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django import forms
from django.urls import path
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from import_export.admin import ImportExportModelAdmin
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget, DateWidget
from django.contrib import messages
from django.shortcuts import redirect, render
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.utils.safestring import mark_safe
from .models import ConfiguracionEmail
from django.core.mail import get_connection
from django.core.mail import EmailMultiAlternatives
from .models import EscanerKardex
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import HistorialPrecio


from .models import (
    OrdenCompra, PartidaCompra, Inventario, DocumentoOrdenCompra,
    CatalogoMedicamento, EstatusProcedimiento, Empresa, SocioComercial,
    Licitacion, PartidaRequerimiento, RegistroUbicacion, Contrato,
    FianzaContrato, ClaveContrato, OrdenSuministro, RemisionEntrega,
    EntradaAlmacen, Almacen, TraspasoIntercompany, PartidaOrden,
    Cotizacion, PartidaCotizacion, PedidoDirecto, DEPENDENCIAS_MAESTRAS, PerfilEquipo
)

@admin.action(description='📥 Descargar seleccionados a Excel (CSV)')
def exportar_a_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reporte_erp.csv"'
    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response)
    campos = [field.name for field in modeladmin.model._meta.fields]
    writer.writerow(campos)
    for objeto in queryset:
        writer.writerow([getattr(objeto, campo) for campo in campos])
    return response

# ==========================================
# 📊 MOTOR UNIFICADO: ANÁLISIS COMERCIAL EJECUTIVO
# ==========================================
@admin.action(description='📊 Descargar Análisis Comercial (Unificado)')
def exportar_analisis_unificado(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    es_individual = queryset.count() == 1
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis Comercial"

    font_cabecera = Font(bold=True, color="000000")
    fill_cabecera = PatternFill("solid", fgColor="D9D9D9")
    fill_datos_grales = PatternFill("solid", fgColor="EFEFEF")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    formato_moneda = '"$"#,##0.00'

    if es_individual:
        obj = queryset.first()
        es_licitacion = hasattr(obj, 'num_procedimiento')
        
        folio = getattr(obj, 'num_procedimiento', getattr(obj, 'folio', 'S/D'))
        cliente = obj.get_dependencia_display() if es_licitacion else (obj.razon_social or obj.get_dependencia_display() or "S/D")
        
        f_pub = obj.fecha_publicacion.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_publicacion else 'N/A'
        f_ape = obj.fecha_apertura.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_apertura else 'N/A'
        f_jun = obj.fecha_junta.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_junta else 'N/A'
        f_fal = obj.fecha_fallo.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_fallo else 'N/A'

        response['Content-Disposition'] = f'attachment; filename="Analisis_Ejecutivo_{folio}.xlsx"'

        ws.append(['', 'NÚMERO DE EXPEDIENTE', 'DEPENDENCIA', 'Observaciones '])
        ws.append(['', folio, cliente, ''])
        ws.append([])
        ws.append(['', 'FECHA DE PUBLICACIÓN', 'JUNTA DE ACLARACIONES', 'APERTURA DE PROPUESTAS', 'ACTO DE FALLO'])
        ws.append(['', f_pub, f_jun, f_ape, f_fal])
        ws.append([])
        ws.append([])

        for r in [1, 4]:
            for c in [2, 3, 4, 5]:
                celda = ws.cell(row=r, column=c)
                celda.font = font_cabecera
                celda.fill = fill_datos_grales
                celda.alignment = alineacion_centro
                celda.border = borde_delgado
        for r in [2, 5]:
            for c in [2, 3, 4, 5]:
                celda = ws.cell(row=r, column=c)
                celda.alignment = alineacion_centro
                celda.border = borde_delgado

        encabezados = ['PARTIDA', 'CLAVE', 'Descripcion (completa) ', 'denominacion_generica', 'denominacion_distintiva', 'Socio Comercial', 'CANTIDAD MÍNIMA', 'CANTIDAD MÁXIMA', 'PRECIO UNITARIO', 'IMPORTE MÁXIMO', 'RESULTADO']
        ws.append(encabezados)
        fila_encabezados = ws.max_row

        for col_num in range(1, len(encabezados) + 1):
            celda = ws.cell(row=fila_encabezados, column=col_num)
            celda.font = font_cabecera
            celda.fill = fill_cabecera
            celda.alignment = alineacion_centro
            celda.border = borde_delgado

        partidas = obj.partidas.all() if es_licitacion else obj.partidas_cotizacion.all()
        for i, p in enumerate(partidas, 1):
            med = p.medicamento
            clave = med.clave_sector if med else 'S/C'
            desc_completa = med.descripcion if med else 'S/D'
            generica = med.denominacion_generica if med else 'S/D'
            distintiva = med.denominacion_distintiva if med and med.denominacion_distintiva else ''
            socio = med.socio_contacto.nombre if med and med.socio_contacto else 'Sin Laboratorio'
            
            cant_min = getattr(p, 'cantidad_minima', 0)
            cant_max = getattr(p, 'cantidad_maxima', getattr(p, 'cantidad', 0))
            precio = float(getattr(p, 'precio', getattr(p, 'precio_unitario', 0)))
            importe_max = cant_max * precio
            
            if es_licitacion:
                resultado = getattr(p, 'resultado', 'Pendiente')
            else:
                resultado = obj.get_estatus_display() if hasattr(obj, 'get_estatus_display') else 'Pendiente'

            ws.append([getattr(p, 'numero_partida', i), clave, desc_completa, generica, distintiva, socio, cant_min, cant_max, precio, importe_max, resultado])
            
            r = ws.max_row
            for c in range(1, len(encabezados) + 1):
                celda = ws.cell(row=r, column=c)
                celda.border = borde_delgado
                celda.alignment = alineacion_centro
                if c in [9, 10]: celda.number_format = formato_moneda
                if c in [3, 4, 5, 6]: celda.alignment = alineacion_izq

    else:
        response['Content-Disposition'] = 'attachment; filename="Analisis_Comercial_Global.xlsx"'
        encabezados = ['FOLIO', 'PARTIDA', 'CLAVE', 'Descripcion (completa) ', 'denominacion_generica', 'denominacion_distintiva', 'Socio Comercial', 'CANTIDAD MÍNIMA', 'CANTIDAD MÁXIMA', 'PRECIO UNITARIO', 'IMPORTE MÁXIMO', 'RESULTADO']
        ws.append(encabezados)
        for col_num in range(1, len(encabezados) + 1):
            celda = ws.cell(row=1, column=col_num)
            celda.font = font_cabecera
            celda.fill = fill_cabecera
            celda.alignment = alineacion_centro
            celda.border = borde_delgado

        for obj in queryset:
            es_licitacion = hasattr(obj, 'num_procedimiento')
            folio = getattr(obj, 'num_procedimiento', getattr(obj, 'folio', 'S/D'))
            partidas = obj.partidas.all() if hasattr(obj, 'num_procedimiento') else obj.partidas_cotizacion.all()
            for i, p in enumerate(partidas, 1):
                med = p.medicamento
                clave = med.clave_sector if med else 'S/C'
                desc_completa = med.descripcion if med else 'S/D'
                generica = med.denominacion_generica if med else 'S/D'
                distintiva = med.denominacion_distintiva if med and med.denominacion_distintiva else ''
                socio = med.socio_contacto.nombre if med and med.socio_contacto else 'Sin Laboratorio'

                cant_min = getattr(p, 'cantidad_minima', 0)
                cant_max = getattr(p, 'cantidad_maxima', getattr(p, 'cantidad', 0))
                precio = float(getattr(p, 'precio', getattr(p, 'precio_unitario', 0)))
                importe_max = cant_max * precio
                
                if es_licitacion:
                    resultado = getattr(p, 'resultado', 'Pendiente')
                else:
                    resultado = obj.get_estatus_display() if hasattr(obj, 'get_estatus_display') else 'Pendiente'

                ws.append([folio, getattr(p, 'numero_partida', i), clave, desc_completa, generica, distintiva, socio, cant_min, cant_max, precio, importe_max, resultado])
                r = ws.max_row
                for c in range(1, len(encabezados) + 1):
                    celda = ws.cell(row=r, column=c)
                    celda.border = borde_delgado
                    celda.alignment = alineacion_centro
                    if c in [10, 11]: celda.number_format = formato_moneda
                    if c in [4, 5, 6, 7]: celda.alignment = alineacion_izq

    wb.save(response)
    return response

# ==========================================
# 🏭 MOTOR UNIFICADO: REPORTE POR SOCIO COMERCIAL
# ==========================================
@admin.action(description='🏭 Descargar Reporte Agrupado por Socios (Unificado)')
def exportar_socios_unificado(modeladmin, request, queryset):
    from django.db.models import Sum
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    es_individual = queryset.count() == 1
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios Comerciales"

    font_cabecera = Font(bold=True, color="000000")
    fill_cabecera = PatternFill("solid", fgColor="D9D9D9")
    font_subtotal = Font(bold=True, italic=True)
    fill_subtotal = PatternFill("solid", fgColor="F2F2F2")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    formato_moneda = '"$"#,##0.00'

    if es_individual:
        obj = queryset.first()
        folio = getattr(obj, 'num_procedimiento', getattr(obj, 'folio', 'S/D'))
        es_licitacion = hasattr(obj, 'num_procedimiento')
        
        f_pub = obj.fecha_publicacion.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_publicacion else 'N/A'
        f_ape = obj.fecha_apertura.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_apertura else 'N/A'
        f_jun = obj.fecha_junta.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_junta else 'N/A'
        f_fal = obj.fecha_fallo.strftime('%d/%m/%Y %H:%M') if es_licitacion and obj.fecha_fallo else 'N/A'

        response['Content-Disposition'] = f'attachment; filename="Reporte_Socios_{folio}.xlsx"'

        ws.append(['', 'NÚMERO DE EXPEDIENTE', 'DEPENDENCIA', 'Observaciones '])
        ws.append(['', folio, obj.get_dependencia_display() if es_licitacion else (obj.razon_social or "S/D"), ''])
        ws.append([])
        ws.append(['', 'FECHA DE PUBLICACIÓN', 'JUNTA DE ACLARACIONES', 'APERTURA DE PROPUESTAS', 'ACTO DE FALLO'])
        ws.append(['', f_pub, f_jun, f_ape, f_fal])
        ws.append([])
        ws.append([])

        for r in [1, 4]:
            for c in [2, 3, 4, 5]:
                celda = ws.cell(row=r, column=c)
                celda.font = font_cabecera
                celda.fill = PatternFill("solid", fgColor="EFEFEF")
                celda.alignment = alineacion_centro
                celda.border = borde_delgado
        for r in [2, 5]:
            for c in [2, 3, 4, 5]:
                celda = ws.cell(row=r, column=c)
                celda.alignment = alineacion_centro
                celda.border = borde_delgado
    else:
        response['Content-Disposition'] = 'attachment; filename="Reporte_Socios_Global.xlsx"'

    encabezados = ['Socio Comercial', 'Partida', 'Clave', 'Descripcion (completa) ', 'denominacion_generica', 'denominacion_distintiva', 'fabricante', 'rfc_fabricante', 'pais_fabricacion', 'num_registro_sanitario', 'num_prorroga', 'codigo_barras', 'fecha_expedicion', 'fecha_vigencia', 'Cantidad solicitada', 'Piezas en inventario', 'Ubicación del Inventario', '$ referencia', 'precio costo ', 'precio unitario', 'Importe']
    ws.append(encabezados)
    fila_encabezados = ws.max_row
    
    for col_num in range(1, len(encabezados) + 1):
        celda = ws.cell(row=fila_encabezados, column=col_num)
        celda.font = font_cabecera
        celda.fill = fill_cabecera
        celda.alignment = alineacion_centro
        celda.border = borde_delgado

    filas = []
    for obj in queryset:
        es_licitacion = hasattr(obj, 'num_procedimiento')
        partidas = obj.partidas.all() if es_licitacion else obj.partidas_cotizacion.all()
        
        for i, p in enumerate(partidas, 1):
            med = p.medicamento
            if not med: continue
            
            socio = med.socio_contacto.nombre if med.socio_contacto else 'Sin socio comercial'
            clave = med.clave_sector
            cant = getattr(p, 'cantidad_maxima', getattr(p, 'cantidad', 0))
            precio = float(getattr(p, 'precio', getattr(p, 'precio_unitario', 0)))
            costo = float(getattr(p, 'costo', 0))
            importe = cant * precio
            num_partida = getattr(p, 'numero_partida', i)
            
            stock_disp = Inventario.objects.filter(medicamento=med, cantidad_disponible__gt=0).select_related('almacen')
            total_piezas = sum(item.cantidad_disponible for item in stock_disp)
            texto_ubic = " | ".join([f"{item.almacen.nombre if item.almacen else 'Bodega'} ({item.cantidad_disponible})" for item in stock_disp]) if stock_disp else "Sin Inventario"

            f_exp = med.fecha_expedicion.strftime('%Y-%m-%d') if med.fecha_expedicion else ''
            f_vig = med.fecha_vigencia.strftime('%Y-%m-%d') if med.fecha_vigencia else ''

            filas.append([
                socio, num_partida, clave, med.descripcion, med.denominacion_generica, 
                med.denominacion_distintiva or '', med.fabricante or '', med.rfc_fabricante or '', 
                med.pais_fabricacion or '', med.num_registro_sanitario or '', med.num_prorroga or '', 
                med.codigo_barras or '', f_exp, f_vig, cant, total_piezas, texto_ubic, 
                0, costo, precio, importe
            ])
            
    filas.sort(key=lambda x: str(x[0])) 
    
    current_fab = None
    subtotal_importe = 0
    
    for fila in filas:
        fab = fila[0]
        if current_fab != fab:
            if current_fab is not None:
                r = ['' for _ in range(21)]
                r[0] = f'Total {current_fab}'
                r[20] = subtotal_importe
                ws.append(r)
                fila_subtotal = ws.max_row
                for c in range(1, 22): 
                    ws.cell(row=fila_subtotal, column=c).fill = fill_subtotal
                    ws.cell(row=fila_subtotal, column=c).font = font_subtotal
                    ws.cell(row=fila_subtotal, column=c).border = borde_delgado
                ws.cell(row=fila_subtotal, column=21).number_format = formato_moneda
            
            current_fab = fab
            subtotal_importe = 0
            mostrar_fab = fab
        else:
            mostrar_fab = ''
            
        subtotal_importe += fila[20] 
        fila[0] = mostrar_fab
        ws.append(fila)
        fila_actual = ws.max_row
        
        for c in range(1, 22):
            celda = ws.cell(row=fila_actual, column=c)
            celda.border = borde_delgado
            if c in [18, 19, 20, 21]: celda.number_format = formato_moneda

    if current_fab is not None:
        r = ['' for _ in range(21)]
        r[0] = f'Total {current_fab}'
        r[20] = subtotal_importe
        ws.append(r)
        fila_subtotal = ws.max_row
        for c in range(1, 22):
            ws.cell(row=fila_subtotal, column=c).fill = fill_subtotal
            ws.cell(row=fila_subtotal, column=c).font = font_subtotal
            ws.cell(row=fila_subtotal, column=c).border = borde_delgado
        ws.cell(row=fila_subtotal, column=21).number_format = formato_moneda

    wb.save(response)
    return response

# ==========================================
# --- CLASES DEL ADMIN ---
# ==========================================
class SocioComercialWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        if not value or str(value).strip() in ['#N/A', 'N/A', 'NA', '']:
            return None
        nombre_socio = str(value).strip()
        socio, created = self.model.objects.get_or_create(
            nombre=nombre_socio,
            defaults={'correos': 'pendiente@actualizar.com'}
        )
        return socio

class CatalogoMedicamentoResource(resources.ModelResource):
    socio_contacto = fields.Field(
        column_name='socio_contacto',
        attribute='socio_contacto',
        widget=SocioComercialWidget(SocioComercial, field='nombre')
    )
    fecha_expedicion = fields.Field(
        column_name='fecha_expedicion',
        attribute='fecha_expedicion',
        widget=DateWidget(format='%Y-%m-%d')
    )
    fecha_vigencia = fields.Field(
        column_name='fecha_vigencia',
        attribute='fecha_vigencia',
        widget=DateWidget(format='%Y-%m-%d')
    )

    class Meta:
        model = CatalogoMedicamento
        fields = (
            'id', 'clave_sector', 'descripcion', 'denominacion_generica', 
            'denominacion_distintiva', 'fabricante', 'socio_contacto', 
            'rfc_fabricante', 'pais_fabricacion', 'num_registro_sanitario', 
            'num_prorroga', 'codigo_barras', 'fecha_expedicion', 'fecha_vigencia'
        )
        
# 👇 ESTA ES LA MAGIA DEL INVENTARIO 👇
    def before_import_row(self, row, **kwargs):
        # 1. Si tipo_producto está vacío
        val_tipo = str(row.get('tipo_producto', '')).strip()
        if not val_tipo or val_tipo.upper() in ['N/A', 'NONE', 'NULL']:
            row['tipo_producto'] = 'COMERCIAL'

        # 2. Si la fecha de caducidad está vacía
        val_fecha = str(row.get('fecha_caducidad', '')).strip()
        if not val_fecha or val_fecha.upper() in ['N/A', 'NONE', 'NULL', 'S/F']:
            row['fecha_caducidad'] = '2030-12-31'


class SemaforoVigenciaFilter(admin.SimpleListFilter):
    title = 'Vigencia de Registro (Semáforo)'
    parameter_name = 'semaforo_vigencia'

    def lookups(self, request, model_admin):
        return (
            ('rojo', '🔴 Vencidos'),
            ('amarillo', '🟡 Por Vencer (≤ 90 días)'),
            ('verde', '🟢 Vigentes (> 90 días)'),
        )

    def queryset(self, request, queryset):
        hoy = timezone.now().date()
        import datetime
        limite_90 = hoy + datetime.timedelta(days=90)
        
        if self.value() == 'rojo':
            return queryset.filter(fecha_vigencia__lt=hoy)
        if self.value() == 'amarillo':
            return queryset.filter(fecha_vigencia__gte=hoy, fecha_vigencia__lte=limite_90)
        if self.value() == 'verde':
            return queryset.filter(fecha_vigencia__gt=limite_90)
        return queryset

class CatalogoMedicamentoAdmin(ImportExportModelAdmin): 
    resource_class = CatalogoMedicamentoResource
    list_per_page = 30
    search_fields = ['clave_sector', 'denominacion_distintiva', 'denominacion_generica', 'socio_contacto__nombre', 'fabricante']
    list_filter = ('socio_contacto', 'fabricante', SemaforoVigenciaFilter)
    
    list_display = (
        'clave_sector', 
        'descripcion_corta', 
        'denominacion_generica', 
        'denominacion_distintiva',
        'socio_contacto', 
        'fabricante', 
        'registro_sanitario_formato',
        'fecha_expedicion',
        'semaforo_vigencia_display' 
    )

    def registro_sanitario_formato(self, obj):
        return obj.num_registro_sanitario
    registro_sanitario_formato.short_description = "Reg. Sanitario"
    
    def descripcion_corta(self, obj):
        if obj.descripcion and len(obj.descripcion) > 35:
            return f"{obj.descripcion[:35]}..."
        return obj.descripcion
    descripcion_corta.short_description = "Descripción"

    def semaforo_vigencia_display(self, obj):
        if not obj.fecha_vigencia:
            return mark_safe('<span style="color: #999;">Sin Fecha</span>') 
            
        hoy = timezone.now().date()
        dias_restantes = (obj.fecha_vigencia - hoy).days
        fecha_str = obj.fecha_vigencia.strftime('%d/%m/%Y')
        
        if dias_restantes < 0:
            return format_html('<span style="background-color: #dc3545; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">🔴 Venció ({})</span>', fecha_str)
        elif dias_restantes <= 90:
            return format_html('<span style="background-color: #f39c12; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">🟡 Por Vencer ({})</span>', fecha_str)
        else:
            return format_html('<span style="color: #28a745; font-weight: bold;">🟢 {}</span>', fecha_str)
            
    semaforo_vigencia_display.short_description = "Vigencia Reg."
    semaforo_vigencia_display.admin_order_field = 'fecha_vigencia'

class RegistroUbicacionAdmin(admin.ModelAdmin):
    list_per_page = 30
    list_display = ('usuario', 'fecha_hora', 'latitud', 'longitud', 'ver_mapa')
    list_filter = ('usuario', 'fecha_hora')
    actions = [exportar_a_csv] 

    def ver_mapa(self, obj):
        if obj.latitud and obj.longitud:
            url = f"https://www.google.com/maps?q={obj.latitud},{obj.longitud}"
            return format_html('<a href="{}" target="_blank" style="background-color: #28a745; color: white; padding: 5px 10px; border-radius: 4px; text-decoration: none;">📍 Ver en Mapa</a>', url)
        return "Sin ubicación"
    ver_mapa.short_description = "Google Maps"

@admin.register(Almacen)
class AlmacenAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'descripcion')
    search_fields = ('nombre',)

class LicitacionForm(forms.ModelForm):
    pegar_excel = forms.CharField(label="📥 Carga Masiva (Pegar desde Excel)", required=False, widget=forms.Textarea(attrs={'rows': 6, 'placeholder': 'Ejemplo:\n1\t010.000.4434.00\tPARACETAMOL 500MG\t500\t1000'}), help_text="Copia de Excel 4 o 5 columnas juntas: PARTIDA | CLAVE | DESCRIPCIÓN | (OPCIONAL: PIEZAS MIN) | PIEZAS MAX.")
    fecha_publicacion = forms.DateTimeField(label="Fecha y hora de publicación", required=False, widget=forms.DateTimeInput(format='%Y-%m-%dT%H:%M', attrs={'type': 'datetime-local'}))
    fecha_apertura = forms.DateTimeField(label="Fecha y hora de apertura", required=False, widget=forms.DateTimeInput(format='%Y-%m-%dT%H:%M', attrs={'type': 'datetime-local'}))
    fecha_junta = forms.DateTimeField(label="Fecha y hora de junta de aclaraciones", required=False, widget=forms.DateTimeInput(format='%Y-%m-%dT%H:%M', attrs={'type': 'datetime-local'}))
    fecha_fallo = forms.DateTimeField(label="Fecha y hora del acto del Fallo", required=False, widget=forms.DateTimeInput(format='%Y-%m-%dT%H:%M', attrs={'type': 'datetime-local'}))
    class Meta:
        model = Licitacion
        fields = '__all__'

class PartidaRequerimientoInline(admin.TabularInline):
    model = PartidaRequerimiento
    extra = 0 
    autocomplete_fields = ['medicamento'] 
    fields = ('numero_partida', 'medicamento', 'cantidad_maxima', 'precio', 'resultado', 'motivo_perdida')
    ordering = ['numero_partida']

    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return ('precio', 'cantidad_maxima', 'licitacion', 'medicamento')
        return super().get_readonly_fields(request, obj)

@admin.register(Licitacion)
class LicitacionAdmin(admin.ModelAdmin):
    list_per_page = 30
    form = LicitacionForm 
    list_display = ('num_procedimiento', 'dependencia', 'estatus_color', 'apertura_semaforo', 'fecha_fallo', 'notificar_whatsapp_btn')
    search_fields = ['num_procedimiento', 'dependencia']
    inlines = [PartidaRequerimientoInline]
    
    actions = [exportar_a_csv, exportar_analisis_unificado, exportar_socios_unificado]
    
    change_form_template = "admin/licitaciones/licitacion/change_form.html"

    def estatus_color(self, obj):
        if not obj.estatus: return "-"
        estado = obj.estatus.estado.upper()
        if estado == 'EN_PROCESO': color_fondo = '#3498db' 
        elif estado == 'ADJUDICADO': color_fondo = '#2ecc71' 
        elif estado == 'PERDIDO': color_fondo = '#e74c3c' 
        else: color_fondo = '#95a5a6' 
        return format_html('<span style="color: white; background-color: {}; padding: 4px 8px; border-radius: 6px; font-weight: bold; font-size: 11px;">{}</span>', color_fondo, estado)
    estatus_color.short_description = 'Estatus'

    def apertura_semaforo(self, obj):
        if not obj.fecha_apertura: 
            return format_html('<span style="color: #aaa;">{}</span>', 'Sin fecha asignada')
            
        hoy = timezone.now().date()
        dias_faltantes = (obj.fecha_apertura.date() - hoy).days
        
        if dias_faltantes < 0:
            return format_html('<span style="color: gray;"><b>Ya pasó</b> (hace {} días)</span>', abs(dias_faltantes))
        elif dias_faltantes <= 3:
            return format_html('<span style="color: red;"><b>Crítico</b> (Faltan {} días)</span>', dias_faltantes)
        elif dias_faltantes <= 7:
            return format_html('<span style="color: orange;"><b>Atención</b> (Faltan {} días)</span>', dias_faltantes)
        else:
            return format_html('<span style="color: green;"><b>A tiempo</b> (Faltan {} días)</span>', dias_faltantes)
            
    apertura_semaforo.short_description = "Fecha de Apertura"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/exportar-analisis/', self.admin_site.admin_view(self.exportar_analisis_view), name='exportar_analisis_licitacion'),
            path('<path:object_id>/exportar-reporte-laboratorios/', self.admin_site.admin_view(self.exportar_reporte_laboratorios_view), name='exportar_reporte_laboratorios_licitacion'),
            path('<path:object_id>/borrar-partidas/', self.admin_site.admin_view(self.borrar_partidas_view), name='borrar_partidas_licitacion'),
            path('<path:object_id>/notificar-socios/', self.admin_site.admin_view(self.notificar_socios_view), name='notificar_socios_licitacion'),
            path('<path:object_id>/notificar-resultados/', self.admin_site.admin_view(self.notificar_resultados_view), name='notificar_resultados_licitacion'),
            path('<path:object_id>/notificar-whatsapp/', self.admin_site.admin_view(self.notificar_whatsapp_view), name='notificar_whatsapp_licitacion'),
        ]
        return custom_urls + urls

    def exportar_analisis_view(self, request, object_id):
        return exportar_analisis_unificado(self, request, Licitacion.objects.filter(id=object_id))

    def exportar_reporte_laboratorios_view(self, request, object_id):
        return exportar_socios_unificado(self, request, Licitacion.objects.filter(id=object_id))

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change) 
        datos_excel = form.cleaned_data.get('pegar_excel')
        if datos_excel:
            import io, csv
            from django.db import transaction
            importes_agregados, claves_nuevas_creadas, errores_filas = 0, [], []
            texto_seguro = datos_excel.replace('\r\n', '\n').replace('\r', '\n')
            f = io.StringIO(texto_seguro)
            lector = csv.reader(f, dialect='excel-tab')
            for columnas in lector:
                if not columnas or not "".join(columnas).strip(): continue
                if len(columnas) >= 4:
                    try:
                        with transaction.atomic():
                            partida_val, clave_val, descripcion_val = columnas[0].strip(), columnas[1].strip(), columnas[2].strip().replace('\n', ' ') 
                            if len(columnas) >= 5 and columnas[4].strip():
                                min_str, max_str = columnas[3].strip().replace(',', '').replace(' ', ''), columnas[4].strip().replace(',', '').replace(' ', '')
                            else:
                                min_str, max_str = "0", columnas[3].strip().replace(',', '').replace(' ', '')
                            if not clave_val or not max_str: continue
                            num_partida = int(float(partida_val))
                            num_piezas_min, num_piezas_max = int(float(min_str)) if min_str else 0, int(float(max_str))
                            medicamento_db = CatalogoMedicamento.objects.filter(clave_sector=clave_val).first()
                            if not medicamento_db:
                                medicamento_db = CatalogoMedicamento(clave_sector=clave_val, descripcion=descripcion_val, denominacion_generica=descripcion_val, fabricante='')
                                medicamento_db.save()
                                claves_nuevas_creadas.append(clave_val)
                            PartidaRequerimiento.objects.create(licitacion=obj, numero_partida=num_partida, medicamento=medicamento_db, cantidad_minima=num_piezas_min, cantidad_maxima=num_piezas_max, costo=0, precio=0)
                            importes_agregados += 1
                    except Exception as e:
                        errores_filas.append(f"Partida {columnas[0]}: Error ({e})")
                        continue 
            if importes_agregados > 0: messages.success(request, f"¡Éxito! Se cargaron {importes_agregados} partidas al evento.")
            if claves_nuevas_creadas: messages.warning(request, f"🔔 Se agregaron {len(claves_nuevas_creadas)} CLAVES NUEVAS.")

    def borrar_partidas_view(self, request, object_id):
        licitacion = self.get_object(request, object_id)
        cantidad = licitacion.partidas.count()
        licitacion.partidas.all().delete()
        messages.success(request, f"Se eliminaron {cantidad} partidas.")
        return redirect('admin:licitaciones_licitacion_change', object_id)

    def notificar_socios_view(self, request, object_id):
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        from django.conf import settings
        
        licitacion = self.get_object(request, object_id)
        partidas = licitacion.partidas.all()
        socios_dict = {}
        
        for p in partidas:
            clave_actual = p.medicamento.clave_sector
            medicamentos_relacionados = CatalogoMedicamento.objects.filter(clave_sector=clave_actual)
            for med in medicamentos_relacionados:
                socio = med.socio_contacto
                if socio:
                    if socio.id not in socios_dict: socios_dict[socio.id] = {'socio': socio, 'partidas': []}
                    if p not in socios_dict[socio.id]['partidas']: socios_dict[socio.id]['partidas'].append(p)

        if request.method == 'POST':
            socios_seleccionados = request.POST.getlist('socios')
            archivos_adjuntos = request.FILES.getlist('adjuntos') 
            empresa_id = request.POST.get('empresa_emisora')
            empresa_emisora = Empresa.objects.get(id=empresa_id)
            
            nombre_empresa_up = empresa_emisora.nombre.upper()
            if "SAGO" in nombre_empresa_up: lista_respuesta = ["sagomedical.licitaciones@gmail.com"]
            elif "GSM" in nombre_empresa_up: lista_respuesta = ["gsm.licitaciones@gmail.com"]
            else: lista_respuesta = ["licitaciones2@gpharma.com"]

            if empresa_emisora.correos_notificacion:
                empleados = [c.strip() for c in empresa_emisora.correos_notificacion.split(',') if c.strip()]
                lista_respuesta.extend(empleados)

            remitente = empresa_emisora.correo_remitente if empresa_emisora.correo_remitente else "notificaciones@jcv-sistemas.lat"
            from_email_str = f'"{empresa_emisora.nombre}" <{remitente}>'
            
            conexion_dinamica = get_connection()
            correos_enviados = 0
            
            for socio_id in socios_seleccionados:
                data = socios_dict.get(int(socio_id))
                if not data: continue
                socio = data['socio']
                
                correos_raw = socio.correos if socio.correos else ""
                destinatarios = [c.strip() for c in correos_raw.split(',') if c.strip()]
                if not destinatarios:
                    messages.warning(request, f"⚠️ Saltado: {socio.nombre} no tiene un correo válido.")
                    continue

                asunto = f"Requerimiento de Cotización - Evento {licitacion.num_procedimiento}"
                
                partidas_html = []
                for p in data['partidas']:
                    partidas_html.append({
                        'partida': p.numero_partida,
                        'clave': p.medicamento.clave_sector,
                        'descripcion': p.medicamento.descripcion,
                        'cantidad': f"{p.cantidad_maxima:,}" 
                    })
                
                if "SAGO" in nombre_empresa_up: color_empresa = "#8B0000"
                elif "GAMS" in nombre_empresa_up: color_empresa = "#005b96"
                elif "GSM" in nombre_empresa_up: color_empresa = "#218838"
                else: color_empresa = "#333333"
                
                contexto_email = {
                    'socio_nombre': socio.nombre,
                    'evento_num': licitacion.num_procedimiento,
                    'dependencia': licitacion.dependencia,
                    'empresa_emisora': empresa_emisora.nombre,
                    'url_logo': empresa_emisora.url_logo if hasattr(empresa_emisora, 'url_logo') and empresa_emisora.url_logo else None,
                    'color_empresa': color_empresa,
                    'fecha_actual': timezone.now().strftime('%d/%m/%Y'),
                    'items': partidas_html,
                }
                
                html_content = render_to_string('admin/licitaciones/licitacion/emails/cotizacion_email.html', contexto_email)
                text_content = strip_tags(html_content) 

                try:
                    correo = EmailMultiAlternatives(
                        subject=asunto, body=text_content, from_email=from_email_str, 
                        to=destinatarios, connection=conexion_dinamica, bcc=lista_respuesta, reply_to=lista_respuesta
                    )
                    correo.attach_alternative(html_content, "text/html")
                    for archivo in archivos_adjuntos: 
                        archivo.seek(0)
                        correo.attach(archivo.name, archivo.read(), archivo.content_type)
                    correo.send(fail_silently=False)
                    correos_enviados += 1
                except Exception as e: 
                    messages.error(request, f"Error al enviar a {socio.nombre}: {e}")
            
            if correos_enviados > 0: messages.success(request, f"¡Se enviaron {correos_enviados} requerimientos!")
            return redirect('admin:licitaciones_licitacion_change', object_id)

        context = {'title': f'Notificar Socios: {licitacion.num_procedimiento}', 'licitacion': licitacion, 'socios_data': socios_dict.values(), 'opts': self.model._meta, 'empresas_grupo': Empresa.objects.all(), 'has_view_permission': self.has_view_permission(request, licitacion)}
        return render(request, 'admin/licitaciones/licitacion/notificar_socios.html', context)

    def notificar_resultados_view(self, request, object_id):
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        from django.conf import settings
        
        licitacion = self.get_object(request, object_id)
        partidas = licitacion.partidas.select_related('medicamento__socio_contacto')
        socios_dict = {}
        
        for p in partidas:
            socio = p.medicamento.socio_contacto
            if socio:
                if socio.id not in socios_dict: socios_dict[socio.id] = {'socio': socio, 'asignadas': [], 'perdidas_precio': [], 'perdidas_tecnica': [], 'pendientes': []}
                if p.resultado == 'Asignada': socios_dict[socio.id]['asignadas'].append(p)
                elif p.resultado == 'Perdida por precio': socios_dict[socio.id]['perdidas_precio'].append(p)
                elif p.resultado == 'Perdida técnicamente': socios_dict[socio.id]['perdidas_tecnica'].append(p)
                else: socios_dict[socio.id]['pendientes'].append(p)

        if request.method == 'POST':
            socios_seleccionados = request.POST.getlist('socios')
            archivos_adjuntos = request.FILES.getlist('adjuntos') 
            empresa_id = request.POST.get('empresa_emisora')
            empresa_emisora = Empresa.objects.get(id=empresa_id)
            
            nombre_empresa_up = empresa_emisora.nombre.upper()
            if "SAGO" in nombre_empresa_up: lista_respuesta = ["sagomedical.licitaciones@gmail.com"]
            elif "GSM" in nombre_empresa_up: lista_respuesta = ["gsm.licitaciones@gmail.com"]
            else: lista_respuesta = ["licitaciones2@gpharma.com"]

            if empresa_emisora.correos_notificacion:
                empleados = [c.strip() for c in empresa_emisora.correos_notificacion.split(',') if c.strip()]
                lista_respuesta.extend(empleados)

            remitente = empresa_emisora.correo_remitente if empresa_emisora.correo_remitente else "notificaciones@jcv-sistemas.lat"
            from_email_str = f'"{empresa_emisora.nombre}" <{remitente}>'

            conexion_dinamica = get_connection()
            correos_enviados = 0
            
            for socio_id in socios_seleccionados:
                data = socios_dict.get(int(socio_id))
                if not data: continue
                socio = data['socio']
                
                correos_raw = socio.correos if socio.correos else ""
                destinatarios = [c.strip() for c in correos_raw.split(',') if c.strip()]
                if not destinatarios:
                    messages.warning(request, f"⚠️ Saltado: {socio.nombre} no tiene un correo válido.")
                    continue

                asunto = f"Resultados Oficiales (Evento {licitacion.num_procedimiento}) - {empresa_emisora.nombre}"
                if "SAGO" in nombre_empresa_up: color_empresa = "#8B0000"
                elif "GAMS" in nombre_empresa_up: color_empresa = "#005b96"
                elif "GSM" in nombre_empresa_up: color_empresa = "#218838"
                else: color_empresa = "#333333"

                def formato_item(p):
                    return {
                        'partida': p.numero_partida,
                        'clave': p.medicamento.clave_sector,
                        'descripcion': p.medicamento.descripcion,
                        'cantidad': f"{p.cantidad_maxima:,}",
                        'motivo': p.motivo_perdida if p.motivo_perdida else "No especificado en el sistema."
                    }

                contexto_email = {
                    'socio_nombre': socio.nombre,
                    'evento_num': licitacion.num_procedimiento,
                    'empresa_emisora': empresa_emisora.nombre,
                    'url_logo': empresa_emisora.url_logo if hasattr(empresa_emisora, 'url_logo') and empresa_emisora.url_logo else None,
                    'color_empresa': color_empresa,
                    'fecha_actual': timezone.now().strftime('%d/%m/%Y'),
                    'url_drive': licitacion.url_carpeta_drive if hasattr(licitacion, 'url_drive') and licitacion.url_carpeta_drive else None,
                    'asignadas': [formato_item(p) for p in data['asignadas']],
                    'perdidas_precio': [formato_item(p) for p in data['perdidas_precio']],
                    'perdidas_tecnica': [formato_item(p) for p in data['perdidas_tecnica']],
                }
                
                html_content = render_to_string('admin/licitaciones/licitacion/emails/resultados_email.html', contexto_email)
                text_content = strip_tags(html_content)

                try:
                    correo = EmailMultiAlternatives(
                        subject=asunto, body=text_content, from_email=from_email_str, 
                        to=destinatarios, connection=conexion_dinamica, bcc=lista_respuesta, reply_to=lista_respuesta
                    )
                    correo.attach_alternative(html_content, "text/html")
                    for archivo in archivos_adjuntos:
                        archivo.seek(0)
                        correo.attach(archivo.name, archivo.read(), archivo.content_type)
                    correo.send(fail_silently=False)
                    correos_enviados += 1
                except Exception as e: 
                    messages.error(request, f"Error al enviar a {socio.nombre}: {e}")
            
            if correos_enviados > 0: messages.success(request, f"Resultados enviados a {correos_enviados} socios.")
            return redirect('admin:licitaciones_licitacion_change', object_id)

        context = {'title': f'Notificar Resultados: {licitacion.num_procedimiento}', 'licitacion': licitacion, 'socios_data': socios_dict.values(), 'opts': self.model._meta, 'empresas_grupo': Empresa.objects.all(), 'has_view_permission': self.has_view_permission(request, licitacion)}
        return render(request, 'admin/licitaciones/licitacion/notificar_resultados.html', context)
    
    def notificar_whatsapp_btn(self, obj):
        tiene_resultados = obj.partidas.exclude(resultado='Pendiente').exists()
        if not tiene_resultados:
            return mark_safe('<span style="color: #94a3b8; background: #f1f5f9; padding: 4px 8px; border-radius: 4px; font-size: 11px; font-weight: 600;"><i class="fas fa-lock"></i> Capturar Resultados</span>')
            
        return format_html('<a class="button" href="{}/notificar-whatsapp/" onclick="window.open(this.href, \'whatsapp_popup\', \'width=500,height=520,resizable=no,scrollbars=yes\'); return false;" style="background-color: #25D366; color:white; padding: 5px 10px; border-radius: 4px; text-decoration: none; font-weight:bold; font-size:11px;"><i class="fab fa-whatsapp"></i> Notificar WA</a>', obj.id)
    notificar_whatsapp_btn.short_description = "WhatsApp"

    def notificar_whatsapp_view(self, request, object_id):
        from twilio.rest import Client
        licitacion = self.get_object(request, object_id)
        equipo = PerfilEquipo.objects.select_related('user').filter(activo=True)

        if request.method == 'POST':
            miembros_ids = request.POST.getlist('miembros')
            estatus_txt = licitacion.estatus.estado if licitacion.estatus else "Finalizado"
            mensaje_texto = f"📋 *GPHARMA ERP - Alerta de Fallo*\n\nEl analista ha registrado los resultados del evento:\n• *Procedimiento:* {licitacion.num_procedimiento}\n• *Dependencia:* {licitacion.get_dependencia_display()}\n• *Resultado General:* {estatus_txt}\n\nPor favor ingresa al ERP para consultar el desglose de las claves asignadas."

            if miembros_ids:
                try:
                    client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
                    enviados = 0
                    for m_id in miembros_ids:
                        miembro = PerfilEquipo.objects.get(id=m_id)
                        if miembro.whatsapp:
                            client.messages.create(from_=settings.TWILIO_PHONE_NUMBER, body=mensaje_texto, to=f"whatsapp:{miembro.whatsapp}")
                            enviados += 1
                    return HttpResponse(f"<html><body><script>alert('¡Éxito! Notificaciones enviadas correctamente a {enviados} integrantes.'); window.close();</script></body></html>")
                except Exception as e:
                    return HttpResponse(f"<html><body><h3 style='color:#dc3545;'>Error al disparar Twilio: {str(e)}</h3><button onclick='window.close()'>Cerrar Ventana</button></body></html>")

        context = {'licitacion': licitacion, 'equipo': equipo, 'opts': self.model._meta}
        return render(request, 'admin/licitaciones/licitacion/notificar_whatsapp_popup.html', context)

@admin.register(Empresa)
class EmpresaAdmin(ImportExportModelAdmin):
    list_per_page = 30
    list_display = ('nombre', 'rfc', 'representante', 'telefono')
    search_fields = ('nombre', 'rfc')
    
    def has_module_permission(self, request):
        return request.user.is_superuser
        
    def has_view_permission(self, request, obj=None):
        return request.user.is_superuser

@admin.register(SocioComercial)
class SocioComercialAdmin(admin.ModelAdmin):
    list_per_page = 30
    list_display = ('nombre', 'correos', 'telefono')
    search_fields = ('nombre', 'correos')

@admin.register(PartidaRequerimiento)
class PartidaRequerimientoAdmin(admin.ModelAdmin):
    list_per_page = 50
    list_display = ('licitacion', 'numero_partida', 'clave_medicamento', 'cantidad_maxima', 'precio', 'importe_total', 'resultado')
    list_filter = ('resultado', 'licitacion__num_procedimiento')
    search_fields = ('numero_partida', 'medicamento__clave_sector', 'medicamento__denominacion_generica', 'licitacion__num_procedimiento')
    ordering = ('licitacion', 'numero_partida')
    list_editable = ('precio', 'resultado')

    def has_module_permission(self, request):
        return False

    def clave_medicamento(self, obj):
        return f"{obj.medicamento.clave_sector} - {obj.medicamento.denominacion_generica[:30]}..."
    clave_medicamento.short_description = "Clave / Medicamento"
    clave_medicamento.admin_order_field = 'medicamento__clave_sector'

    def importe_total(self, obj):
        total = (obj.cantidad_maxima or 0) * (obj.precio or 0)
        return format_html('<b>${}</b>', f"{total:,.2f}")
    importe_total.short_description = "Importe Máximo"
    importe_total.admin_order_field = 'cantidad_maxima'

admin.site.register(CatalogoMedicamento, CatalogoMedicamentoAdmin)
admin.site.register(RegistroUbicacion, RegistroUbicacionAdmin)

# ==========================================
# --- FASE 2: MÓDULO DE CONTRATOS Y LOGÍSTICA ---
# ==========================================
class FianzaContratoInline(admin.TabularInline):
    model = FianzaContrato
    extra = 1

class ClaveContratoInline(admin.TabularInline):
    model = ClaveContrato
    extra = 0
    autocomplete_fields = ['medicamento']
    fields = ('medicamento', 'cantidad_minima', 'cantidad_maxima', 'precio_neto', 'piezas_historicas_solicitadas', 'piezas_historicas_entregadas', 'importe_maximo')
    readonly_fields = ('importe_maximo',)

    def importe_maximo(self, obj):
        if obj.cantidad_maxima and obj.precio_neto:
            total = obj.cantidad_maxima * obj.precio_neto
            return format_html('<b>${}</b>', f"{total:,.2f}")
        return "$0.00"
    importe_maximo.short_description = "Importe Máx."

class EmpresaSeguraWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        if value: value = str(value).strip()
        return super().clean(value, row, **kwargs)

class LicitacionOpcionalWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        if not value or str(value).strip() in ['', 'None', 'N/A', 'NA']: return None
        return self.model.objects.filter(num_procedimiento=str(value).strip()).first()

class CargaMaestraContratoResource(resources.ModelResource):
    empresa = fields.Field(column_name='empresa', attribute='empresa', widget=EmpresaSeguraWidget(Empresa, field='nombre'))
    licitacion = fields.Field(column_name='licitacion', attribute='licitacion_origen', widget=LicitacionOpcionalWidget(Licitacion, field='num_procedimiento'))
    fecha_inicio = fields.Field(column_name='fecha_inicio', attribute='fecha_inicio', widget=DateWidget(format='%Y-%m-%d'))
    fecha_fin = fields.Field(column_name='fecha_fin', attribute='fecha_fin', widget=DateWidget(format='%Y-%m-%d'))

    class Meta:
        model = Contrato
        fields = ('numero_contrato', 'dependencia', 'empresa', 'fecha_inicio', 'fecha_fin', 'licitacion')
        import_id_fields = ('numero_contrato',)
        skip_unchanged = False

    def before_import_row(self, row, **kwargs):
        for key in list(row.keys()):
            if isinstance(row[key], str):
                row[key] = row[key].strip()
                
        for key in list(row.keys()):
            k_lower = str(key).strip().lower()
            val = row[key]
            
            if k_lower == 'empresa': row[key] = str(val).strip() if val else None
            elif k_lower == 'licitacion': row[key] = str(val).strip() if val else None
            elif k_lower in ['fecha_inicio', 'fecha_fin']:
                if str(val).strip() in ['', 'None', 'N/A', 'NA']: row[key] = None

        lic_str = None
        dep_str = 'S/D'
        for key, val in row.items():
            k_lower = str(key).strip().lower()
            if k_lower == 'licitacion':
                lic_str = str(val).strip() if val else None
            if k_lower == 'dependencia':
                dep_str = str(val).strip() if val else 'S/D'

        if lic_str:
            lic_obj = Licitacion.objects.filter(num_procedimiento=lic_str).first()
            if not lic_obj:
                Licitacion.objects.bulk_create([
                    Licitacion(num_procedimiento=lic_str, dependencia=dep_str)
                ])

        clave_sec = None
        for key, val in row.items():
            if str(key).strip().lower() == 'clave_sector':
                clave_sec = str(val).strip() if val else None
                break

        if clave_sec:
            medicamento = CatalogoMedicamento.objects.filter(clave_sector=clave_sec).first()
            if not medicamento:
                CatalogoMedicamento.objects.create(
                    clave_sector=clave_sec,
                    descripcion='Agregado por carga de Contrato', 
                    denominacion_generica='PENDIENTE ASIGNAR MARCA'
                )

    def after_import_row(self, row, row_result, **kwargs):
        if kwargs.get('dry_run'):
            return 
            
        num_contrato = None
        clave_sec = None
        cant_min, cant_max, precio = 0, 0, 0.0
        hist_sol, hist_ent = 0, 0
        
        for key, val in row.items():
            k_lower = str(key).strip().lower()
            if k_lower == 'numero_contrato': num_contrato = str(val).strip() if val else None
            elif k_lower == 'clave_sector': clave_sec = str(val).strip() if val else None
            elif k_lower == 'cantidad_minima':
                try: cant_min = int(float(val or 0))
                except: cant_min = 0
            elif k_lower == 'cantidad_maxima':
                try: cant_max = int(float(val or 0))
                except: cant_max = 0
            elif k_lower == 'precio_neto':
                try: precio = float(val or 0.0)
                except: precio = 0.0
            elif k_lower == 'piezas_historicas_solicitadas':
                try: hist_sol = int(float(val or 0))
                except: hist_sol = 0
            elif k_lower == 'piezas_historicas_entregadas':
                try: hist_ent = int(float(val or 0))
                except: hist_ent = 0
                
        if num_contrato and clave_sec:
            contrato_obj = Contrato.objects.filter(numero_contrato=num_contrato).first()
            med_obj = CatalogoMedicamento.objects.filter(clave_sector=clave_sec).first()
            
            if contrato_obj and med_obj:
                ClaveContrato.objects.update_or_create(
                    contrato=contrato_obj,
                    medicamento=med_obj,
                    defaults={
                        'cantidad_minima': cant_min,
                        'cantidad_maxima': cant_max,
                        'precio_neto': precio,
                        'piezas_historicas_solicitadas': hist_sol,
                        'piezas_historicas_entregadas': hist_ent
                    }
                )

@admin.register(Contrato)
class ContratoAdmin(ImportExportModelAdmin): 
    resource_class = CargaMaestraContratoResource 
    
    list_per_page = 30
    list_display = ('numero_contrato',  'licitacion_origen', 'dependencia', 'empresa', 'fecha_inicio', 'fecha_fin', 'fianzas_badges', 'mostrar_monto', 'mostrar_avance', 'mostrar_abasto', 'boton_expediente', 'monto_penalizado')
    search_fields = ('numero_contrato', 'dependencia', 'fianzas__numero_fianza', 'licitacion_origen__num_procedimiento') 
    list_filter = ('empresa', 'fecha_fin', 'licitacion_origen')
    inlines = [FianzaContratoInline, ClaveContratoInline]
    autocomplete_fields = ['licitacion_origen']
    change_form_template = "admin/licitaciones/licitacion/boton_contrato.html"

    fieldsets = (
        ('🏢 Datos Generales del Contrato', {'fields': ('empresa', 'numero_contrato', 'dependencia')}),
        ('📅 Vigencia', {'fields': ('fecha_inicio', 'fecha_fin') }),
        ('🤖 Vinculación', {'fields': ('licitacion_origen',), 'description': 'Selecciona la Licitación de origen.'}),
        ('📂 Expediente Físico/Servidor', {'fields': ('ruta_carpeta_servidor',), 'description': 'Pega la ruta de la carpeta compartida de tu red.'}),
    )

    def fianzas_badges(self, obj):
        fianzas = obj.fianzas.all()
        if not fianzas: return format_html('<span style="color: #ccc;">{}</span>', 'Sin Fianzas')
        badges = []
        for f in fianzas:
            badges.append(f'<span style="background-color: #17a2b8; color: white; padding: 3px 8px; border-radius: 10px; font-size: 10px; display: inline-block; margin-bottom: 3px;">{f.get_tipo_display()}: {f.numero_fianza}</span>')
        return format_html(''.join(badges))
    fianzas_badges.short_description = "Pólizas / Fianzas"

    def mostrar_monto(self, obj):
        return format_html('<b style="color: #28a745;">${}</b>', f"{obj.monto_total_contrato:,.2f}")
    mostrar_monto.short_description = "Monto Total Máximo"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('<path:object_id>/seleccionar-claves/', self.admin_site.admin_view(self.seleccionar_claves_view), name='seleccionar_claves_contrato')]
        return custom_urls + urls
    
    def boton_expediente(self, obj):
        if hasattr(obj, 'ruta_carpeta_servidor') and obj.ruta_carpeta_servidor:
            ruta_segura = obj.ruta_carpeta_servidor.replace('\\', '\\\\')
            return format_html('<button type="button" onclick="navigator.clipboard.writeText(\'{}\'); alert(\'¡Ruta copiada!\');" style="background-color: #6c757d; color: white; padding: 4px 10px; border: none; border-radius: 4px; font-size: 11px; font-weight: bold; cursor: pointer;">📋 Copiar Ruta</button>', ruta_segura)
        return mark_safe('<span style="color: #ccc;">-</span>')
    boton_expediente.short_description = "Expediente"

    def mostrar_avance(self, obj):
        avance = obj.porcentaje_avance
        color = 'green' if avance < 80 else ('orange' if avance <= 95 else 'red')
        return format_html('<b style="color: {};">{}%</b>', color, avance)
    mostrar_avance.short_description = 'Avance Financiero'

    def mostrar_abasto(self, obj):
        abasto = obj.porcentaje_abasto
        color = 'green' if abasto >= 95 else ('orange' if abasto >= 80 else 'red')
        return format_html('<b style="color: {};">{}%</b>', color, abasto)
    mostrar_abasto.short_description = 'Nivel de Abasto'

    def seleccionar_claves_view(self, request, object_id):
        contrato = self.get_object(request, object_id)
        if not contrato.licitacion_origen:
            messages.error(request, "Este contrato no tiene una Licitación de origen vinculada.")
            return redirect('admin:licitaciones_contrato_change', object_id)

        claves_ya_en_contrato = contrato.claves.values_list('medicamento_id', flat=True)
        partidas_ganadas = contrato.licitacion_origen.partidas.filter(resultado='Asignada').exclude(medicamento_id__in=claves_ya_en_contrato)

        if request.method == 'POST':
            partidas_seleccionadas = request.POST.getlist('partidas')
            claves_creadas = 0
            for partida_id in partidas_seleccionadas:
                partida = PartidaRequerimiento.objects.get(id=partida_id)
                ClaveContrato.objects.create(
                    contrato=contrato, medicamento=partida.medicamento,
                    precio_neto=partida.precio if partida.precio else 0,
                    cantidad_minima=partida.cantidad_minima if partida.cantidad_minima else 0,
                    cantidad_maxima=partida.cantidad_maxima if partida.cantidad_maxima else 0
                )
                claves_creadas += 1
            if claves_creadas > 0: messages.success(request, f"¡Éxito! Se agregaron {claves_creadas} claves al contrato.")
            return redirect('admin:licitaciones_contrato_change', object_id)

        context = {'title': f'Elegir Claves Ganadas: {contrato.numero_contrato}', 'contrato': contrato, 'partidas': partidas_ganadas, 'opts': self.model._meta, 'has_view_permission': self.has_view_permission(request, contrato)}
        return render(request, 'admin/licitaciones/licitacion/seleccionar_claves_contrato.html', context)

    def monto_penalizado(self, obj):
        ordenes = OrdenSuministro.objects.filter(partidas__clave_contrato__contrato=obj).distinct()
        try: total_multas = sum(float(orden.penalizacion_estimada) for orden in ordenes)
        except (ValueError, TypeError): total_multas = 0.0
            
        if total_multas > 0: return format_html('<span style="color: #dc3545; font-weight: bold;">- ${}</span>', f"{total_multas:,.2f}")
        return mark_safe('<span style="color: #28a745; font-weight: bold;">$0.00</span>')
    monto_penalizado.short_description = "Penalizaciones"

# ==========================================
# RECURSOS LOGÍSTICOS
# ==========================================
class ClaveContratoWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        num_contrato = str(row.get('numero_contrato', '')).strip()
        clave_sec = str(row.get('clave_medicamento', '')).strip()
        if num_contrato and clave_sec:
            return ClaveContrato.objects.filter(contrato__numero_contrato__iexact=num_contrato, medicamento__clave_sector__iexact=clave_sec).first()
        return None

class OrdenSuministroResource(resources.ModelResource):
    actions = ['marcar_como_canceladas']
    clave_contrato = fields.Field(column_name='clave_medicamento', attribute='clave_contrato', widget=ClaveContratoWidget(ClaveContrato, field='id'))
    clave_medicamento_historico = fields.Field(column_name='clave_medicamento', attribute='clave_medicamento_historico')
    numero_contrato_historico = fields.Field(column_name='numero_contrato', attribute='numero_contrato_historico')
    fecha_recepcion = fields.Field(column_name='fecha_recepcion', attribute='fecha_recepcion', widget=DateWidget(format='%Y-%m-%d'))
    fecha_limite = fields.Field(column_name='fecha_limite', attribute='fecha_limite', widget=DateWidget(format='%Y-%m-%d'))
    fecha_entrega_real = fields.Field(column_name='fecha_entrega_real', attribute='fecha_entrega_real', widget=DateWidget(format='%Y-%m-%d'))

    class Meta:
        model = OrdenSuministro
        fields = ('razon_social', 'numero_orden_suministro', 'numero_procedimiento_extra', 'numero_contrato_historico', 'clave_medicamento_historico', 'descripcion_medicamento', 'cantidad_solicitada', 'cantidad_entregada', 'precio_unitario', 'clues_destino', 'entidad_destino', 'nombre_unidad', 'fecha_recepcion', 'fecha_limite', 'fecha_entrega_real', 'clave_contrato', 'estatus')
        import_id_fields = ('numero_orden_suministro', 'clave_medicamento_historico')

    def before_import_row(self, row, **kwargs):
        for key in list(row.keys()):
            if isinstance(row[key], str): row[key] = row[key].strip()
        row['numero_contrato_historico'] = str(row.get('numero_contrato', '')).strip()
        row['clave_medicamento_historico'] = str(row.get('clave_medicamento', '')).strip()

        try:
            solicitada = int(float(row.get('cantidad_solicitada', 0) or 0))
            entregada = int(float(row.get('cantidad_entregada', 0) or 0))
        except (ValueError, TypeError):
            solicitada, entregada = 0, 0

        if entregada > 0:
            if entregada >= solicitada: row['estatus'] = 'ENTREGADA'
            else: row['estatus'] = 'PARCIAL'
        else: row['estatus'] = 'PENDIENTE'

class PartidaOrdenInline(admin.TabularInline):
    model = PartidaOrden
    extra = 1
    autocomplete_fields = ['medicamento']
    fields = ('clave_contrato', 'medicamento', 'cantidad_solicitada', 'precio_unitario', 'cantidad_entregada')

@admin.register(OrdenSuministro)
class OrdenSuministroAdmin(ImportExportModelAdmin):
    resource_class = OrdenSuministroResource
    inlines = [PartidaOrdenInline]
    list_per_page = 50
    list_display = ('razon_social', 'numero_orden_suministro', 'dependencia', 'nombre_unidad', 'piezas_solicitadas', 'piezas_entregadas', 'piezas_pendientes', 'estatus_logistico', 'monto_penalizacion', 'estatus', 'btn_surtir')
    search_fields = ('numero_orden_suministro', 'dependencia', 'razon_social', 'nombre_unidad')
    list_filter = ('razon_social', 'dependencia', 'estatus')
    list_editable = ('estatus',)

    fieldsets = (
        ('1. Datos del Documento', {'fields': ('tipo_documento', 'numero_orden_suministro', 'fecha_recepcion', 'fecha_limite')}),
        ('2. Cliente / Dependencia', {'fields': ('razon_social', 'dependencia', 'entidad_destino', 'clues_destino', 'nombre_unidad')}),
        ('3. Logística y Estatus', {'fields': ('estatus', 'fecha_entrega_real', 'motivo_incidencia')}),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).filter(tipo_documento='SUMINISTRO')

    def piezas_solicitadas(self, obj): return sum(p.cantidad_solicitada for p in obj.partidas.all())
    piezas_solicitadas.short_description = "Cant. Solicitada"

    def piezas_entregadas(self, obj):
        enviadas = sum((r.cantidad_entregada or 0) for r in obj.remisiones.exclude(estatus_viaje='RECHAZO'))
        total_solicitado = sum(p.cantidad_solicitada for p in obj.partidas.all())
        if enviadas >= total_solicitado and total_solicitado > 0: return format_html('<b style="color: #28a745;">{}</b>', f"{enviadas:,}") 
        return format_html('<b style="color: #007bff;">{}</b>', f"{enviadas:,}")
    piezas_entregadas.short_description = "Cant. Entregada"

    def piezas_pendientes(self, obj):
        enviadas = sum((r.cantidad_entregada or 0) for r in obj.remisiones.exclude(estatus_viaje='RECHAZO'))
        total_solicitado = sum(p.cantidad_solicitada for p in obj.partidas.all())
        pendientes = total_solicitado - enviadas
        if pendientes <= 0: return mark_safe('<b style="color: #28a745;">0</b>') 
        return format_html('<b style="color: #dc3545;">{}</b>', f"{pendientes:,}")
    piezas_pendientes.short_description = "Cant. Pendiente"

    def estatus_logistico(self, obj):
        color = "#6c757d"
        texto = obj.get_estatus_display()
        if obj.estatus == 'ENTREGADA': color = "#28a745"
        elif obj.estatus == 'PENDIENTE': color = "#ffc107"
        elif obj.estatus == 'DEVUELTA': color = "#dc3545"
        elif obj.estatus == 'PARCIAL': color = "#17a2b8"
        elif obj.estatus == 'NO_ATENDIDA': color = "#343a40"

        if obj.estatus in ['CANCELADA', 'CANCELADA_EVIDENCIA']:
            piezas_entregadas = sum((r.cantidad_entregada or 0) for r in obj.remisiones.exclude(estatus_viaje='RECHAZO'))
            if piezas_entregadas > 0:
                return format_html('<div style="text-align: center; line-height: 1.2;"><span style="background-color: #000; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 10px; display: block; margin-bottom: 2px;">🚨 CANCELADA (TIENE EVIDENCIA)</span><span style="font-size: 10px; color: #dc3545; font-weight: bold;">Se entregaron {} pzas</span></div>', f"{piezas_entregadas:,}")
            return format_html('<span style="background-color: #6f42c1; color: white; padding: 5px 10px; border-radius: 4px; font-weight: bold; font-size: 11px;">🚫 CANCELADA (Sin entregas)</span>')

        if obj.estatus in ['PENDIENTE', 'PARCIAL'] and obj.dias_atraso > 0:
            return format_html('<span style="background-color: #dc3545; color: white; padding: 5px 10px; border-radius: 4px; font-weight: bold; font-size: 11px;">⚠️ ATRASADA ({} DÍAS)</span>', obj.dias_atraso)

        return format_html('<span style="background-color: {}; color: white; padding: 5px 10px; border-radius: 4px; font-weight: bold; font-size: 11px;">{}</span>', color, texto)
    estatus_logistico.short_description = "Semáforo"

    def monto_penalizacion(self, obj):
        try: penalizacion = float(obj.penalizacion_estimada)
        except (ValueError, TypeError): penalizacion = 0.0
        if penalizacion > 0: return format_html('<span style="color: #dc3545; font-weight: bold;">- ${}</span>', f"{penalizacion:,.2f}")
        return mark_safe('<span style="color: #ccc;">$0.00</span>')
    monto_penalizacion.short_description = "Penalización Acumulada"
    
    def btn_surtir(self, obj):
        if obj.estatus == 'ENTREGADA': return mark_safe('<span style="color: green; font-weight:bold;">✔ Completada</span>')
        return format_html('<a class="button" href="{}/surtir/" style="background-color: #007bff; color:white; padding: 5px 10px; border-radius: 4px; text-decoration: none; font-weight:bold;">🚚 Surtir</a>', obj.id)
    btn_surtir.short_description = "Logística"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('<path:object_id>/surtir/', self.admin_site.admin_view(self.surtir_orden_view), name='surtir_orden_logistica')]
        return custom_urls + urls

    def surtir_orden_view(self, request, object_id):
        orden = self.get_object(request, object_id)
        claves_buscar = [p.clave_contrato.medicamento.clave_sector for p in orden.partidas.all() if p.clave_contrato and p.clave_contrato.medicamento]
        lotes_disponibles = Inventario.objects.filter(medicamento__clave_sector__in=claves_buscar, cantidad_disponible__gt=0).order_by('fecha_caducidad')
        total_solicitado = sum(p.cantidad_solicitada for p in orden.partidas.all())
        ya_enviado = sum(r.cantidad_entregada for r in orden.remisiones.all())
        falta_enviar = total_solicitado - ya_enviado

        if request.method == 'POST':
            lote_id, cantidad_a_enviar, folio_remision = request.POST.get('lote_id'), int(request.POST.get('cantidad', 0)), request.POST.get('folio_remision')
            if cantidad_a_enviar > 0 and lote_id and folio_remision:
                lote_seleccionado = Inventario.objects.get(id=lote_id)
                if cantidad_a_enviar > lote_seleccionado.cantidad_disponible:
                    messages.error(request, "No puedes enviar más piezas de las que hay en ese lote.")
                else:
                    lote_seleccionado.cantidad_disponible -= cantidad_a_enviar
                    lote_seleccionado.save()
                    RemisionEntrega.objects.create(orden=orden, folio_remision_factura=folio_remision, cantidad_entregada=cantidad_a_enviar, lote=lote_seleccionado.lote, caducidad=lote_seleccionado.fecha_caducidad)
                    orden.estatus = 'EN_RUTA'
                    orden.save()
                    messages.success(request, f"¡Camión en ruta! Se despacharon {cantidad_a_enviar} piezas con la remisión {folio_remision}.")
                    return redirect('admin:licitaciones_ordensuministro_changelist')

        context = {'title': f'Surtir Orden: {orden.numero_orden_suministro}', 'orden': orden, 'lotes': lotes_disponibles, 'falta_enviar': falta_enviar, 'opts': self.model._meta}
        return render(request, 'admin/licitaciones/ordensuministro/surtir_orden.html', context)

    @admin.action(description='🚫 Marcar seleccionadas como CANCELADAS')
    def marcar_como_canceladas(self, request, queryset):
        ordenes_procesadas, con_evidencia = 0, 0
        for orden in queryset:
            piezas_entregadas = sum((r.cantidad_entregada or 0) for r in orden.remisiones.exclude(estatus_viaje='RECHAZO'))
            if piezas_entregadas > 0:
                orden.estatus = 'CANCELADA_EVIDENCIA'
                con_evidencia += 1
            else:
                orden.estatus = 'CANCELADA'
            orden.save()
            ordenes_procesadas += 1
        if con_evidencia > 0: messages.warning(request, f"¡Atención! Se cancelaron {ordenes_procesadas} órdenes, {con_evidencia} CON EVIDENCIA.")
        else: messages.success(request, f"Se cancelaron {ordenes_procesadas} órdenes correctamente.")

    def get_actions(self, request):
        actions = super().get_actions(request)
        if not request.user.is_superuser and 'marcar_como_canceladas' in actions: del actions['marcar_como_canceladas']
        return actions

# ==========================================
# REGISTRO DEL NUEVO MÓDULO DE INVENTARIO
# ==========================================

# 1. Traductores Inteligentes para limpiar basura del Excel
class AlmacenSeguroWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        val = str(value).strip() if value else ''
        if val:
            almacen, _ = self.model.objects.get_or_create(nombre=val)
            return almacen
        return None

class MedicamentoSeguroWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        clave = str(value).strip() if value else ''
        if clave:
            # Usamos filter().first() para agarrar la primera y evitar el error de los 7 clones
            med = self.model.objects.filter(clave_sector=clave).first()
            if not med:
                desc = str(row.get('DESCRIPCIÓN', 'Agregado por Excel')).strip()
                med = self.model.objects.create(
                    clave_sector=clave,
                    descripcion=desc,
                    denominacion_generica=desc
                )
            return med
        return None

# 2. El "Mapa" (Resource) principal
class InventarioResource(resources.ModelResource):
    almacen = fields.Field(
        column_name='almacen__nombre',
        attribute='almacen',
        widget=AlmacenSeguroWidget(Almacen, 'nombre')
    )
    medicamento = fields.Field(
        column_name='CLAVE SECTOR',
        attribute='medicamento',
        widget=MedicamentoSeguroWidget(CatalogoMedicamento, 'clave_sector')
    )
    
    # Campos visuales que solo sirven al descargar
    descripcion = fields.Field(attribute='medicamento__denominacion_generica', column_name='DESCRIPCIÓN', readonly=True)
    fabricante = fields.Field(attribute='medicamento__fabricante', column_name='FABRICANTE', readonly=True)
    socio_comercial = fields.Field(attribute='medicamento__socio_contacto__nombre', column_name='SOCIO COMERCIAL', readonly=True)

    class Meta:
        model = Inventario
        fields = ('almacen', 'medicamento', 'descripcion', 'socio_comercial', 'fabricante', 'tipo_producto', 'lote', 'fecha_caducidad', 'cantidad_disponible')
        export_order = fields
        import_id_fields = ('medicamento', 'lote')

    def before_import_row(self, row, **kwargs):
        # 1. Limpiar y forzar el tipo de producto
        val_tipo = str(row.get('tipo_producto', '')).strip().upper()
        if 'GEN' in val_tipo:
            row['tipo_producto'] = 'GENERICO'
        else:
            row['tipo_producto'] = 'COMERCIAL'

        # 2. Rellenar Fechas de Caducidad vacías
        val_fecha = str(row.get('fecha_caducidad', '')).strip()
        if not val_fecha or val_fecha.upper() in ['N/A', 'NONE', 'NULL', 'S/F']:
            row['fecha_caducidad'] = '2030-12-31'
            
# ==========================================
# --- REPORTE DE INVENTARIO "CHULO" ---
# ==========================================
@admin.action(description='📊 Descargar Inventario Ejecutivo (Excel Especial)')
def exportar_inventario_personalizado(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventario_Detallado_Gpharma.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Actual"

    font_cabecera = Font(bold=True, color="000000")
    fill_cabecera = PatternFill("solid", fgColor="D9D9D9")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    encabezados = ['ALMACÉN', 'CLAVE SECTOR', 'DESCRIPCIÓN', 'SOCIO COMERCIAL', 'FABRICANTE', 'TIPO', 'LOTE', 'CADUCIDAD', 'EXISTENCIA']
    ws.append(encabezados)

    for col_num, header in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.font = font_cabecera
        celda.fill = fill_cabecera
        celda.alignment = alineacion_centro
        celda.border = borde_delgado

    for item in queryset:
        med = item.medicamento
        socio = med.socio_contacto.nombre if med.socio_contacto else "SIN ASIGNAR"
        
        fila = [
            item.almacen.nombre if item.almacen else "BODEGA CENTRAL",
            med.clave_sector,
            med.denominacion_generica,
            socio,
            med.fabricante,
            item.get_tipo_producto_display(),
            item.lote,
            item.fecha_caducidad.strftime('%d/%m/%Y') if item.fecha_caducidad else "S/F",
            item.cantidad_disponible
        ]
        ws.append(fila)
        
        r = ws.max_row
        for c in range(1, 10):
            celda = ws.cell(row=r, column=c)
            celda.border = borde_delgado
            celda.alignment = alineacion_centro if c != 3 else alineacion_izq

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12

    wb.save(response)
    return response

@admin.register(Inventario)
class InventarioAdmin(ImportExportModelAdmin):
    resource_class = InventarioResource
    list_per_page = 50
    list_display = ('almacen', 'medicamento', 'tipo_producto', 'lote', 'fecha_caducidad', 'cantidad_disponible', 'fecha_ingreso')
    search_fields = ('medicamento__clave_sector', 'medicamento__denominacion_generica', 'lote')
    list_filter = ('tipo_producto', 'almacen', 'fecha_caducidad',)
    autocomplete_fields = ['medicamento']
    actions = [exportar_inventario_personalizado]

@admin.register(RemisionEntrega)
class RemisionEntregaAdmin(admin.ModelAdmin):
    list_per_page = 50
    list_display = ('folio_remision_factura', 'orden_vinculada', 'cantidad_entregada', 'estatus_viaje', 'imprimir_pdf')
    search_fields = ('folio_remision_factura', 'lote', 'orden__numero_orden_suministro')
    list_filter = ('estatus_viaje',)

    fieldsets = (
        ('📦 Datos del Viaje', {'fields': ('orden', 'folio_remision_factura', 'cantidad_entregada', 'lote', 'caducidad')}),
        ('🚚 Logística y Comprobación', {'fields': ('estatus_viaje', 'archivo_evidencia')}),
        ('🔴 Incidencias: Rechazo', {'fields': ('motivo_rechazo', 'evidencia_rechazo')}),
    )

    def orden_vinculada(self, obj): return obj.orden.numero_orden_suministro
    orden_vinculada.short_description = "OPM Vinculada"

    def imprimir_pdf(self, obj):
        return format_html('<a class="button" href="{}/pdf/" style="background-color: #e74c3c; color:white; padding: 5px 10px; border-radius: 4px; text-decoration: none; font-weight:bold;"><i class="fas fa-file-pdf"></i> Generar PDF</a>', obj.id)
    imprimir_pdf.short_description = "Formato Oficial"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('<path:object_id>/pdf/', self.admin_site.admin_view(self.generar_remision_pdf_view), name='generar_remision_pdf')]
        return custom_urls + urls

    def generar_remision_pdf_view(self, request, object_id):
        from django.template.loader import render_to_string
        import io
        try: from xhtml2pdf import pisa
        except ImportError:
            messages.error(request, "Falta instalar xhtml2pdf. Ejecuta: pip install xhtml2pdf")
            return redirect('admin:licitaciones_remisionentrega_changelist')

        remision = self.get_object(request, object_id)
        orden = remision.orden
        
        num_contrato_hist, clave_hist, desc_hist, precio_hist = getattr(orden, 'numero_contrato_historico', None), getattr(orden, 'clave_medicamento_historico', None), getattr(orden, 'descripcion_medicamento', None), getattr(orden, 'precio_unitario', "0.00")
        
        contrato_vinculado, empresa_emisora = None, None
        primera_partida = orden.partidas.first()
        if primera_partida and hasattr(primera_partida, 'clave_contrato') and primera_partida.clave_contrato:
            contrato_vinculado = primera_partida.clave_contrato.contrato
        if not contrato_vinculado and num_contrato_hist:
            contrato_vinculado = Contrato.objects.filter(numero_contrato__iexact=num_contrato_hist).first()

        if contrato_vinculado:
            empresa_emisora, numero_contrato_final = contrato_vinculado.empresa, contrato_vinculado.numero_contrato
        else: numero_contrato_final = num_contrato_hist or "S/D"

        cliente_final = getattr(orden, 'dependencia', '') or getattr(orden, 'razon_social', '') or "S/D"
        cliente_upper = cliente_final.upper()
        if "SAGO" in cliente_upper or "GAMS" in cliente_upper or "GSM" in cliente_upper:
            cliente_final = getattr(orden, 'entidad_destino', '') or getattr(orden, 'nombre_unidad', '') or "DEPENDENCIA NO ESPECIFICADA"
        dependencia_str = (getattr(orden, 'dependencia', '') or '').upper()

        contexto = {
            'remision': remision, 'orden': orden, 'empresa': empresa_emisora, 'numero_contrato_final': numero_contrato_final,
            'cliente_final': cliente_final, 'dependencia_str': dependencia_str, 'clave_hist': clave_hist, 'desc_hist': desc_hist,
            'precio_hist': precio_hist, 'fecha_actual': getattr(remision, 'fecha_despacho', timezone.now().date()),
        }

        html_string = render_to_string('admin/licitaciones/remisionentrega/pdf/formato_remision.html', contexto)
        result_file = io.BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=result_file)
        
        if not pisa_status.err:
            response = HttpResponse(result_file.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="Remision_{dependencia_str}_{remision.folio_remision_factura}.pdf".replace(" ", "_")'
            return response
        messages.error(request, "Error al procesar el documento PDF.")
        return redirect('admin:licitaciones_remisionentrega_changelist')

# ==========================================
# 🚀 MÓDULO: COTIZACIONES Y VENTAS DIRECTAS
# ==========================================
class CotizacionForm(forms.ModelForm):
    pegar_excel = forms.CharField(label="📥 Carga Masiva (Pegar desde Excel)", required=False, widget=forms.Textarea(attrs={'rows': 6, 'placeholder': 'Ejemplo:\n1\t010.000.4434.00\tPARACETAMOL 500MG\t500\t1000'}))
    fecha_apertura = forms.DateTimeField(label="Fecha y hora de apertura", required=False, widget=forms.DateTimeInput(format='%Y-%m-%dT%H:%M', attrs={'type': 'datetime-local'}))
    class Meta: model = Cotizacion; fields = '__all__'

class PartidaCotizacionInline(admin.TabularInline):
    model = PartidaCotizacion
    extra = 0
    autocomplete_fields = ['medicamento']  
    fields = ('medicamento', 'cantidad', 'precio_unitario', 'importe_visual')
    readonly_fields = ('importe_visual',)

    def importe_visual(self, obj):
        if obj.cantidad and obj.precio_unitario: return format_html('<b>${}</b>', "{:,.2f}".format(float(obj.cantidad) * float(obj.precio_unitario)))
        return "$0.00"
    importe_visual.short_description = "Importe"

@admin.register(Cotizacion)
class CotizacionAdmin(admin.ModelAdmin):
    list_per_page = 30
    form = CotizacionForm
    inlines = [PartidaCotizacionInline]
    change_form_template = "admin/licitaciones/cotizacion/change_form.html"
    list_display = ('folio', 'tipo_procedimiento', 'cliente_visual', 'fecha_emision', 'fecha_apertura', 'total_cotizado', 'estatus_badge', 'btn_convertir')
    search_fields = ('folio', 'razon_social', 'dependencia')
    list_filter = ('tipo_procedimiento', 'estatus', 'fecha_emision')
    actions = [exportar_analisis_unificado, exportar_socios_unificado]

    fieldsets = (
        ('📄 Datos del Evento', {'fields': ('tipo_procedimiento', 'empresa', 'folio', 'fecha_emision', 'fecha_apertura', 'vigencia_dias')}),
        ('🏢 Cliente', {'fields': ('razon_social', 'dependencia')}),
        ('🚦 Estatus', {'fields': ('estatus',)}),
        ('⚡ Carga Automática (Excel)', {'fields': ('pegar_excel',)}),
    )

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change) 
        datos_excel = form.cleaned_data.get('pegar_excel')
        if datos_excel:
            import io, csv
            from django.db import transaction
            importes_agregados, claves_nuevas = 0, []
            f = io.StringIO(datos_excel.replace('\r\n', '\n').replace('\r', '\n'))
            for columnas in csv.reader(f, dialect='excel-tab'):
                if not columnas or not "".join(columnas).strip() or len(columnas) < 4: continue
                try:
                    with transaction.atomic():
                        clave_val, descripcion_val = columnas[1].strip(), columnas[2].strip().replace('\n', ' ') 
                        max_str = columnas[4].strip().replace(',', '').replace(' ', '') if len(columnas) >= 5 and columnas[4].strip() else columnas[3].strip().replace(',', '').replace(' ', '')
                        if not clave_val or not max_str: continue
                        cantidad_final = int(float(max_str))
                        medicamento_db, created = CatalogoMedicamento.objects.get_or_create(clave_sector=clave_val, defaults={'descripcion': descripcion_val, 'denominacion_generica': descripcion_val, 'fabricante': ''})
                        if created: claves_nuevas.append(clave_val)
                        PartidaCotizacion.objects.create(cotizacion=obj, medicamento=medicamento_db, cantidad=cantidad_final, precio_unitario=0.00)
                        importes_agregados += 1
                except Exception: continue 
            if importes_agregados > 0: messages.success(request, f"¡Éxito! Se cargaron {importes_agregados} partidas.")
            if claves_nuevas: messages.warning(request, f"🔔 Se crearon {len(claves_nuevas)} CLAVES NUEVAS.")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/convertir-pedido/', self.admin_site.admin_view(self.convertir_pedido_view), name='convertir_cotizacion_pedido'),
            path('<path:object_id>/exportar-analisis/', self.admin_site.admin_view(self.exportar_analisis_view), name='exportar_analisis_cotizacion'),
            path('<path:object_id>/exportar-reporte-laboratorios/', self.admin_site.admin_view(self.exportar_reporte_laboratorios_view), name='exportar_reporte_laboratorios_cotizacion'),
            path('<path:object_id>/notificar-socios/', self.admin_site.admin_view(self.notificar_socios_view), name='notificar_socios_cotizacion'),
            path('<path:object_id>/notificar-resultados/', self.admin_site.admin_view(self.notificar_resultados_view), name='notificar_resultados_cotizacion'),
        ]
        return custom_urls + urls

    def exportar_analisis_view(self, request, object_id):
        return exportar_analisis_unificado(self, request, Cotizacion.objects.filter(id=object_id))

    def exportar_reporte_laboratorios_view(self, request, object_id):
        return exportar_socios_unificado(self, request, Cotizacion.objects.filter(id=object_id))

    def notificar_socios_view(self, request, object_id):
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        
        try: cotizacion = Cotizacion.objects.get(id=object_id)
        except Cotizacion.DoesNotExist: return redirect('admin:licitaciones_cotizacion_changelist')
            
        socios_dict = {}
        for p in cotizacion.partidas_cotizacion.all():
            socio = p.medicamento.socio_contacto if p.medicamento else None
            if socio:
                if socio.id not in socios_dict: socios_dict[socio.id] = {'socio': socio, 'partidas': []}
                socios_dict[socio.id]['partidas'].append(p)

        if request.method == 'POST':
            socios_seleccionados = request.POST.getlist('socios')
            archivos_adjuntos = request.FILES.getlist('adjuntos')
            empresa_emisora = Empresa.objects.get(id=request.POST.get('empresa_emisora'))
            
            nombre_empresa_up = empresa_emisora.nombre.upper()
            lista_respuesta = ["sagomedical.licitaciones@gmail.com"] if "SAGO" in nombre_empresa_up else (["gsm.licitaciones@gmail.com"] if "GSM" in nombre_empresa_up else ["licitaciones2@gpharma.com"])
            if empresa_emisora.correos_notificacion: lista_respuesta.extend([c.strip() for c in empresa_emisora.correos_notificacion.split(',') if c.strip()])

            remitente = empresa_emisora.correo_remitente if empresa_emisora.correo_remitente else "notificaciones@jcv-sistemas.lat"
            from_email_str = f'"{empresa_emisora.nombre}" <{remitente}>'
            conexion_dinamica = get_connection()
            correos_enviados = 0
            
            for s_id in socios_seleccionados:
                data = socios_dict.get(int(s_id))
                if not data: continue
                socio = data['socio']
                destinatarios = [c.strip() for c in (socio.correos or "").split(',') if c.strip()]
                if not destinatarios: continue
                
                color_empresa = "#8B0000" if "SAGO" in nombre_empresa_up else ("#005b96" if "GAMS" in nombre_empresa_up else ("#218838" if "GSM" in nombre_empresa_up else "#3498db"))
                items = [{'partida': i+1, 'clave': p.medicamento.clave_sector, 'descripcion': p.medicamento.denominacion_generica, 'cantidad': f"{p.cantidad:,}"} for i, p in enumerate(data['partidas'])]
                ctx = {
                    'socio_nombre': socio.nombre, 'evento_num': cotizacion.folio, 'dependencia': self.cliente_visual(cotizacion), 'empresa_emisora': empresa_emisora.nombre,
                    'url_logo': empresa_emisora.url_logo if hasattr(empresa_emisora, 'url_logo') else None, 'color_empresa': color_empresa, 'fecha_actual': timezone.now().strftime('%d/%m/%Y'), 'items': items
                }
                
                html_content = render_to_string('admin/licitaciones/licitacion/emails/cotizacion_email.html', ctx)
                try: 
                    msg = EmailMultiAlternatives(subject=f"Requerimiento de Cotización - Evento {cotizacion.folio}", body=strip_tags(html_content), from_email=from_email_str, to=destinatarios, connection=conexion_dinamica, bcc=lista_respuesta, reply_to=lista_respuesta)
                    msg.attach_alternative(html_content, "text/html")
                    for archivo in archivos_adjuntos: 
                        archivo.seek(0)
                        msg.attach(archivo.name, archivo.read(), archivo.content_type)
                    msg.send(fail_silently=False)
                    correos_enviados += 1
                except Exception as e: messages.error(request, f"Error SMTP con {socio.nombre}: {e}")
                
            if correos_enviados > 0: messages.success(request, f"¡Se enviaron {correos_enviados} requerimientos a laboratorios!")
            return redirect('admin:licitaciones_cotizacion_change', object_id)

        context = {'title': f'Notificar Socios: {cotizacion.folio}', 'evento': cotizacion, 'socios_data': socios_dict.values(), 'opts': self.model._meta, 'empresas_grupo': Empresa.objects.all(), 'has_view_permission': True}
        return render(request, 'admin/licitaciones/cotizacion/notificar_socios_cot.html', context)

    def notificar_resultados_view(self, request, object_id):
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        try: cotizacion = Cotizacion.objects.get(id=object_id)
        except Cotizacion.DoesNotExist: return redirect('admin:licitaciones_cotizacion_changelist')
            
        socios_dict = {}
        for p in cotizacion.partidas_cotizacion.all():
            socio = p.medicamento.socio_contacto if p.medicamento else None
            if socio:
                if socio.id not in socios_dict: socios_dict[socio.id] = {'socio': socio, 'asignadas': [], 'perdidas_precio': [], 'perdidas_tecnica': [], 'pendientes': []}
                if cotizacion.estatus == 'GANADA': socios_dict[socio.id]['asignadas'].append(p)
                elif cotizacion.estatus in ['PERDIDA', 'CANCELADA']: socios_dict[socio.id]['perdidas_precio'].append(p)
                else: socios_dict[socio.id]['pendientes'].append(p)

        if request.method == 'POST':
            socios_seleccionados = request.POST.getlist('socios')
            archivos_adjuntos = request.FILES.getlist('adjuntos')
            empresa_emisora = Empresa.objects.get(id=request.POST.get('empresa_emisora'))
            
            nombre_empresa_up = empresa_emisora.nombre.upper()
            lista_respuesta = ["sagomedical.licitaciones@gmail.com"] if "SAGO" in nombre_empresa_up else (["gsm.licitaciones@gmail.com"] if "GSM" in nombre_empresa_up else ["licitaciones2@gpharma.com"])
            if empresa_emisora.correos_notificacion: lista_respuesta.extend([c.strip() for c in empresa_emisora.correos_notificacion.split(',') if c.strip()])

            remitente = empresa_emisora.correo_remitente if empresa_emisora.correo_remitente else "notificaciones@jcv-sistemas.lat"
            from_email_str = f'"{empresa_emisora.nombre}" <{remitente}>'
            conexion_dinamica = get_connection()
            correos_enviados = 0
            
            for s_id in socios_seleccionados:
                data = socios_dict.get(int(s_id))
                if not data: continue
                socio = data['socio']
                
                destinatarios = [c.strip() for c in (socio.correos or "").split(',') if c.strip()]
                if not destinatarios: continue
                
                color_empresa = "#8B0000" if "SAGO" in nombre_empresa_up else ("#005b96" if "GAMS" in nombre_empresa_up else ("#218838" if "GSM" in nombre_empresa_up else "#3498db"))
                def formato_item(p): return {'partida': '-', 'clave': p.medicamento.clave_sector, 'descripcion': p.medicamento.denominacion_generica, 'cantidad': f"{p.cantidad:,}", 'motivo': f"Cotización global: {cotizacion.get_estatus_display()}"}

                ctx = {
                    'socio_nombre': socio.nombre, 'evento_num': cotizacion.folio, 'empresa_emisora': empresa_emisora.nombre, 'url_logo': empresa_emisora.url_logo if hasattr(empresa_emisora, 'url_logo') else None,
                    'color_empresa': color_empresa, 'fecha_actual': timezone.now().strftime('%d/%m/%Y'), 'url_drive': None, 
                    'asignadas': [formato_item(p) for p in data['asignadas']], 'perdidas_precio': [formato_item(p) for p in data['perdidas_precio']], 'perdidas_tecnica': [formato_item(p) for p in data['perdidas_tecnica']],
                }
                
                html_content = render_to_string('admin/licitaciones/licitacion/emails/resultados_email.html', ctx)
                try: 
                    msg = EmailMultiAlternatives(subject=f"Resultados de Cotización (Evento {cotizacion.folio}) - {empresa_emisora.nombre}", body=strip_tags(html_content), from_email=from_email_str, to=destinatarios, connection=conexion_dinamica, bcc=lista_respuesta, reply_to=lista_respuesta)
                    msg.attach_alternative(html_content, "text/html")
                    for archivo in archivos_adjuntos:
                        archivo.seek(0)
                        msg.attach(archivo.name, archivo.read(), archivo.content_type)
                    msg.send(fail_silently=False)
                    correos_enviados += 1
                except Exception as e: messages.error(request, f"Error SMTP con {socio.nombre}: {e}")
                
            if correos_enviados > 0: messages.success(request, f"¡Resultados enviados a {correos_enviados} laboratorios!")
            return redirect('admin:licitaciones_cotizacion_change', object_id)

        context = {'title': f'Notificar Resultados: {cotizacion.folio}', 'evento': cotizacion, 'socios_data': socios_dict.values(), 'opts': self.model._meta, 'empresas_grupo': Empresa.objects.all(), 'has_view_permission': True}
        return render(request, 'admin/licitaciones/cotizacion/notificar_resultados_cot.html', context)

    def cliente_visual(self, obj): return obj.razon_social if obj.tipo_procedimiento == 'COTIZACION_PRIVADA' else obj.get_dependencia_display()
    cliente_visual.short_description = "Cliente / Dependencia"

    def total_cotizado(self, obj): return format_html('<b style="color: #5e35b1;">${}</b>', "{:,.2f}".format(float(obj.total_cotizacion)))
    total_cotizado.short_description = "Monto Total"

    def estatus_badge(self, obj):
        colores = {'BORRADOR': '#6c757d', 'ENVIADA': '#17a2b8', 'GANADA': '#28a745', 'PERDIDA': '#dc3545', 'CANCELADA': '#343a40'}
        return format_html('<span style="background-color: {}; color: white; padding: 4px 10px; border-radius: 6px; font-weight: bold; font-size: 11px;">{}</span>', colores.get(obj.estatus, '#000'), obj.get_estatus_display())
    estatus_badge.short_description = "Estatus"

    def btn_convertir(self, obj):
        if obj.estatus == 'GANADA': return mark_safe('<span style="color: #28a745; font-weight:bold;">✔ Ya es Pedido</span>')
        if obj.estatus in ['PERDIDA', 'CANCELADA']: return mark_safe('<span style="color: #dc3545; font-weight:bold;">🚫 Rechazada</span>')
        return format_html('<a class="button" href="{}/convertir-pedido/" style="background-color: #28a745; color:white; padding: 5px 10px; border-radius: 4px; text-decoration: none; font-weight:bold;">✨ Hacer Pedido</a>', obj.id)
    btn_convertir.short_description = "Acción"

    def convertir_pedido_view(self, request, object_id):
        cotizacion = self.get_object(request, object_id)
        nueva_orden = PedidoDirecto.objects.create(numero_orden_suministro=f"PED-{cotizacion.folio}", razon_social=cotizacion.razon_social, dependencia=cotizacion.dependencia, fecha_recepcion=timezone.now().date(), estatus='PENDIENTE')
        for partida in cotizacion.partidas_cotizacion.all():
            np = PartidaOrden.objects.create(orden=nueva_orden, cantidad_solicitada=partida.cantidad, precio_unitario=partida.precio_unitario)
            if partida.medicamento: np.medicamento = partida.medicamento; np.save()
            else: np.clave_historica = partida.medicamento.clave_sector; np.save()
        cotizacion.estatus = 'GANADA'
        cotizacion.save()
        messages.success(request, "Se generó el Pedido Directo.")
        return redirect('admin:licitaciones_pedidodirecto_change', nueva_orden.id)

# ==========================================
# 🛑 PANEL DE CUARENTENA, MERMAS Y DEVOLUCIONES
# ==========================================
from .models import IncidenciaInventario

@admin.register(IncidenciaInventario)
class IncidenciaInventarioAdmin(ImportExportModelAdmin):
    list_display = ('id', 'medicamento_clave', 'lote', 'cantidad_afectada', 'motivo', 'resolucion_badge', 'socio_comercial', 'total_recuperado')
    list_filter = ('motivo', 'resolucion', 'socio_comercial', 'genera_nota_credito')
    search_fields = ('medicamento__clave_sector', 'lote', 'observaciones')
    autocomplete_fields = ['inventario_origen', 'medicamento', 'socio_comercial']
    
    fieldsets = (
        ('1. Origen del Problema (Descuento de Stock)', {'fields': ('inventario_origen', 'cantidad_afectada', 'motivo', 'observaciones'), 'description': 'Al guardar, estas piezas se restarán del Inventario sano y pasarán a estado de cuarentena.'}),
        ('2. Resolución y Destino', {'fields': ('resolucion',)}),
        ('3. Finanzas y Reclamaciones (Socio Comercial)', {'fields': ('socio_comercial', 'genera_nota_credito', 'monto_nota_credito', 'aplica_penalizacion', 'monto_penalizacion'), 'description': 'Llena esto si el producto se devolverá al proveedor y se le cobrará mediante nota de crédito o multa.'}),
    )

    def medicamento_clave(self, obj): return obj.medicamento.clave_sector
    medicamento_clave.short_description = "Clave"

    def resolucion_badge(self, obj):
        colores = {'EN_CUARENTENA': '#f39c12', 'DONATIVO': '#17a2b8', 'DEVOLUCION': '#28a745', 'DESTRUCCION': '#dc3545'}
        return format_html('<span style="background-color: {}; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">{}</span>', colores.get(obj.resolucion, '#000'), obj.get_resolucion_display())
    resolucion_badge.short_description = "Estatus / Destino"

    def total_recuperado(self, obj):
        total = float(obj.monto_nota_credito) + float(obj.monto_penalizacion)
        if total > 0: return format_html('<b style="color: #28a745;">+ ${}</b>', f"{total:,.2f}")
        return mark_safe('<span style="color: #ccc;">$0.00</span>') 
    total_recuperado.short_description = "Cobro a Proveedor"

# ==========================================
# 👥 PANEL DE EQUIPO Y ROLES
# ==========================================
@admin.register(PerfilEquipo)
class PerfilEquipoAdmin(admin.ModelAdmin):
    list_display = ('user', 'rol', 'whatsapp', 'activo')
    list_filter = ('rol', 'activo')

# ==========================================
# 🚀 MÓDULO RÁPIDO: PEDIDOS DIRECTOS (10 DÍAS)
# ==========================================
@admin.register(PedidoDirecto)
class PedidoDirectoAdmin(admin.ModelAdmin):
    list_display = ('numero_orden_suministro', 'cliente_info', 'fecha_recepcion', 'fecha_limite', 'semaforo_urgencia', 'estatus')
    list_filter = ('estatus', 'fecha_recepcion')
    search_fields = ('numero_orden_suministro', 'razon_social', 'dependencia')
    exclude = ('tipo_documento',) 

    def get_queryset(self, request):
        return super().get_queryset(request).filter(tipo_documento='PEDIDO')

    def cliente_info(self, obj): return obj.razon_social or obj.get_dependencia_display() or "Sin Cliente"
    cliente_info.short_description = 'Cliente / Dependencia'

    def semaforo_urgencia(self, obj):
        if obj.estatus in ['ENTREGADA', 'CANCELADA', 'CANCELADA_EVIDENCIA', 'DEVUELTA']: return format_html('<span style="color: #94a3b8; font-weight: bold;"><i class="fas fa-check-circle"></i> Cerrado</span>')
        if not obj.fecha_limite: return "-"
        dias_restantes = (obj.fecha_limite - timezone.now().date()).days
        if dias_restantes < 0: return format_html('<span style="color: white; background: #e74c3c; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">🚨 VENCIDO ({} días)</span>', abs(dias_restantes))
        elif dias_restantes <= 3: return format_html('<span style="color: #856404; background: #f1c40f; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">⏳ Crítico ({} días)</span>', dias_restantes)
        else: return format_html('<span style="color: white; background: #2ecc71; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px;">🟢 A tiempo ({} días)</span>', dias_restantes)
    semaforo_urgencia.short_description = 'Tiempo de Entrega'

class ClaveContratoResource(resources.ModelResource):
    contrato = fields.Field(column_name='numero_contrato', attribute='contrato', widget=ForeignKeyWidget(Contrato, field='numero_contrato'))
    medicamento = fields.Field(column_name='clave_sector', attribute='medicamento', widget=ForeignKeyWidget(CatalogoMedicamento, field='clave_sector'))

    class Meta:
        model = ClaveContrato
        fields = ('contrato', 'medicamento', 'cantidad_minima', 'cantidad_maxima', 'precio_neto')
        import_id_fields = ('contrato', 'medicamento')


# =========================================================================
# 🛒 RECUPERADO: MÓDULO DE COMPRAS Y RECEPCIÓN CEGUECIDA
# =========================================================================

class PartidaCompraInline(admin.TabularInline):
    model = PartidaCompra
    extra = 1
    autocomplete_fields = ['medicamento']
    fields = ('medicamento', 'custom_cantidad_pedida', 'precio_unitario', 'importe_visual', 'cantidad_recibida')
    readonly_fields = ('importe_visual', 'custom_cantidad_pedida')

    def custom_cantidad_pedida(self, obj):
        return f"{obj.cantidad:,}" if obj.cantidad else "0"
    custom_cantidad_pedida.short_description = "Cantidad Solicitada"

    def importe_visual(self, obj):
        if obj.cantidad and obj.precio_unitario:
            total = float(obj.cantidad) * float(obj.precio_unitario)
            return format_html('<b>${}</b>', f"{total:,.2f}")
        return "$0.00"
    importe_visual.short_description = "Importe total"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        formfield = super().formfield_for_foreignkey(db_field, request, **kwargs)
        if db_field.name == "medicamento":
            formfield.widget.attrs.update({'style': 'width: 450px; min-width: 450px;'})
        return formfield


class DocumentoOrdenCompraInline(admin.TabularInline):
    model = DocumentoOrdenCompra
    extra = 1
    fields = ('descripcion', 'archivo')


@admin.register(OrdenCompra)
class OrdenCompraAdmin(admin.ModelAdmin):
    list_display = ('folio', 'proveedor', 'fecha_entrega_esperada', 'mostrar_total', 'penalizacion_calculada', 'auditoria_ciega', 'estatus_badge', 'enviar_oc_link')
    list_filter = ('estatus', 'empresa_compradora', 'proveedor')
    search_fields = ('folio', 'proveedor__nombre', 'destino') 
    inlines = [PartidaCompraInline, DocumentoOrdenCompraInline]
    autocomplete_fields = ['proveedor'] 
    list_per_page = 30
    actions = ['marcar_recibida_y_crear_inventario']

    def penalizacion_calculada(self, obj):
        return f"${obj.penalizacion_calculada:,.2f}"
    penalizacion_calculada.short_description = "📉 Penalización por Atraso"

    def enviar_oc_link(self, obj):
        return format_html('<a class="button" href="{}" style="background-color: #5e35b1; color:white; padding: 5px 10px; border-radius: 4px; text-decoration: none; font-size:11px; font-weight:bold;">✉️ Enviar OC</a>', f"{obj.id}/notificar-proveedor/")
    enviar_oc_link.short_description = "Notificar"

    def mostrar_total(self, obj):
        return format_html('<b style="color: #5e35b1;">${}</b>', f"{obj.total_compra:,.2f}")
    mostrar_total.short_description = "Total OC"

    def estatus_badge(self, obj):
        colores = {'BORRADOR': '#6c757d', 'AUTORIZADA': '#007bff', 'TRANSITO': '#fd7e14', 'RECIBIDA': '#28a745', 'CANCELADA': '#dc3545'}
        return format_html('<span style="background-color: {}; color: white; padding: 4px 10px; border-radius: 6px; font-weight: bold; font-size: 11px;">{}</span>', colores.get(obj.estatus, '#000'), obj.get_estatus_display())
    estatus_badge.short_description = "Estatus"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/notificar-proveedor/', self.admin_site.admin_view(self.notificar_proveedor_view), name='notificar_proveedor_oc'),
            path('<path:object_id>/vista-previa-pdf/', self.admin_site.admin_view(self.vista_previa_pdf_view), name='vista_previa_pdf_oc'),
        ]
        return custom_urls + urls

    def vista_previa_pdf_view(self, request, object_id):
        from io import BytesIO
        try: from xhtml2pdf import pisa
        except ImportError:
            messages.error(request, "Falta xhtml2pdf.")
            return redirect('admin:licitaciones_ordencompra_changelist')
        oc = self.get_object(request, object_id)
        contexto = {'oc': oc, 'proveedor': oc.proveedor, 'empresa': oc.empresa_compradora, 'partidas': oc.partidas_compra.all()}
        html_pdf = render_to_string('admin/licitaciones/ordencompra/pdf/orden_compra_pdf.html', contexto)
        result_file = BytesIO()
        pisa.CreatePDF(html_pdf, dest=result_file)
        response = HttpResponse(result_file.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Borrador_{oc.folio}.pdf"'
        return response

    def notificar_proveedor_view(self, request, object_id):
        from io import BytesIO
        try: from xhtml2pdf import pisa
        except ImportError: return redirect('admin:licitaciones_ordencompra_changelist')
        
        oc = self.get_object(request, object_id)
        proveedor = oc.proveedor
        
        if request.method == 'POST':
            empresa = oc.empresa_compradora
            remitente = empresa.correo_remitente if empresa.correo_remitente else "notificaciones@jcv-sistemas.lat"
            destinatarios = [c.strip() for c in proveedor.correos.split(',') if c.strip()]
            
            asunto = f"ORDEN DE COMPRA OFICIAL: {oc.folio} - {empresa.nombre}"
            contexto = {'oc': oc, 'proveedor': proveedor, 'empresa': empresa, 'partidas': oc.partidas_compra.all(), 'total': f"{oc.total_compra:,.2f}"}
            
            html_content = render_to_string('admin/licitaciones/ordencompra/emails/orden_compra_email.html', contexto)
            msg = EmailMultiAlternatives(subject=asunto, body=strip_tags(html_content), from_email=f'"{empresa.nombre}" <{remitente}>', to=destinatarios)
            msg.attach_alternative(html_content, "text/html")
            
            result_file = BytesIO()
            pisa.CreatePDF(render_to_string('admin/licitaciones/ordencompra/pdf/orden_compra_pdf.html', contexto), dest=result_file)
            msg.attach(f"OC_{oc.folio}.pdf", result_file.getvalue(), 'application/pdf')
            msg.send()
            
            if oc.estatus == 'BORRADOR':
                oc.estatus = 'AUTORIZADA'
                oc.save()
            messages.success(request, "🚀 ¡Orden de Compra enviada con éxito!")
            return redirect('admin:licitaciones_ordencompra_changelist')
            
        return render(request, 'admin/licitaciones/ordencompra/confirmar_envio_oc.html', {'oc': oc, 'proveedor': proveedor, 'opts': self.model._meta})

    @admin.action(description='📦 Marcar como RECIBIDA e Ingresar al Inventario')
    def marcar_recibida_y_crear_inventario(self, request, queryset):
        import datetime
        procesadas = 0
        for orden in queryset:
            if orden.estatus == 'RECIBIDA': continue
            orden.estatus = 'RECIBIDA'
            orden.save()
            for p in orden.partidas_compra.all():
                p.custom_cantidad_recibida = p.cantidad
                p.save()
                Inventario.objects.create(
                    medicamento=p.medicamento, cantidad_disponible=p.custom_cantidad_recibida,
                    lote=f"LOT-{orden.folio}", fecha_ingreso=datetime.date.today(),
                    fecha_caducidad=datetime.date.today() + datetime.timedelta(days=730)
                )
            procesadas += 1
        messages.success(request, f"Se procesaron {procesadas} órdenes correctamente.")

    def auditoria_ciega(self, obj):
        reclamado = sum(p.cantidad_recibida for p in obj.partidas_compra.all() if p.cantidad_recibida)
        recibido = sum(e.cantidad_recibida for e in obj.entradaalmacen_set.all() if e.cantidad_recibida)
        if reclamado == 0 and recibido == 0: return mark_safe('<span style="color: #95a5a6; font-weight: bold;">➖ Pendiente</span>')
        if reclamado == recibido: return format_html('<span style="color: #28a745; font-weight: bold;">✅ CUADRA PERFECTO ({})</span>', f"{recibido:,}")
        return format_html('<span style="color: #dc3545; font-weight: bold;">❌ DISCREPANCIA (Comp: {} | Alm: {})</span>', f"{reclamado:,}", f"{recibido:,}")
    auditoria_ciega.short_description = "Auditoría de Recepción"


# =========================================================================
# 📦 RECUPERADO: FORMULARIO DINÁMICO AJAX Y RECEPCIÓN ALMACÉN
# =========================================================================

class EntradaAlmacenForm(forms.ModelForm):
    class Meta:
        model = EntradaAlmacen
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk and self.instance.orden_id:
            self.fields['medicamento'].queryset = CatalogoMedicamento.objects.filter(id__in=PartidaCompra.objects.filter(orden=self.instance.orden).values('medicamento'))
        elif self.data.get('orden'):
            self.fields['medicamento'].queryset = CatalogoMedicamento.objects.filter(id__in=PartidaCompra.objects.filter(orden_id=self.data.get('orden')).values('medicamento'))
        else:
            self.fields['medicamento'].queryset = CatalogoMedicamento.objects.none()

        js = """
        <script type="text/javascript">
            document.addEventListener('DOMContentLoaded', function() {
                var $ = django.jQuery || window.jQuery;
                $('#id_orden').on('change', function() {
                    var ordenId = $(this).val();
                    var $medSelect = $('#id_medicamento');
                    if (!ordenId) { $medSelect.empty(); return; }
                    $medSelect.html('<option value="">🔄 Buscando en la OC...</option>');
                    $.ajax({
                        url: '/admin/licitaciones/entradaalmacen/ajax/load-medicamentos/?orden_id=' + ordenId,
                        success: function(data) {
                            $medSelect.empty().append('<option value="">--- Selecciona clave ---</option>');
                            $.each(data, function(i, item) { $medSelect.append('<option value="'+item.id+'">'+item.text+'</option>'); });
                        }
                    });
                });
            });
        </script>
        """
        if 'orden' in self.fields:
            self.fields['orden'].help_text = mark_safe((self.fields['orden'].help_text or '') + js)


@admin.register(EntradaAlmacen)
class EntradaAlmacenAdmin(admin.ModelAdmin):
    form = EntradaAlmacenForm
    list_per_page = 30
    list_display = ('orden', 'almacen_destino', 'medicamento', 'custom_cant_recibida', 'lote', 'ver_acuse', 'ver_factura', 'documentacion_ok', 'fecha_ingreso')
    search_fields = ('orden__folio', 'medicamento__clave_sector', 'lote')
    list_filter = ('almacen_destino', 'documentacion_completa', 'fecha_ingreso')
    autocomplete_fields = ['orden']

    def custom_cant_recibida(self, obj):
        return f"{obj.cantidad_recibida:,}" if obj.cantidad_recibida else "0"
    custom_cant_recibida.short_description = "Cantidad Recibida"

    fieldsets = (
        ('🔗 Vinculación', {'fields': ('almacen_destino', 'orden', 'medicamento')}),
        ('📦 Datos Físicos', {'fields': ('cantidad_recibida', 'lote', 'fecha_caducidad')}),
        ('📋 Calidad y Logística', {'fields': ('documentacion_completa', 'ubicacion', 'observaciones_calidad', 'acuse_recibo', 'factura_proveedor')}),
    )

    def get_urls(self):
        return [path('ajax/load-medicamentos/', self.admin_site.admin_view(self.load_medicamentos), name='ajax_load_medicamentos_almacen')] + super().get_urls()

    def load_medicamentos(self, request):
        from django.http import JsonResponse
        orden_id = request.GET.get('orden_id')
        if not orden_id: return JsonResponse([])
        partidas = PartidaCompra.objects.filter(orden_id=orden_id).select_related('medicamento')
        return JsonResponse([{'id': p.medicamento.id, 'text': f"Clave: {p.medicamento.clave_sector} | {p.medicamento.denominacion_generica[:40]}..."} for p in partidas], safe=False)

    def ver_acuse(self, obj):
        if obj.acuse_recibo: return format_html('<a href="{}" target="_blank" style="background-color: #17a2b8; color: white; padding: 4px 8px; border-radius: 4px; text-decoration: none; font-size: 11px; font-weight: bold;">📄 Acuse</a>', obj.acuse_recibo.url)
        return "-"
    ver_acuse.short_description = "Acuse"

    def ver_factura(self, obj):
        if obj.factura_proveedor: return format_html('<a href="{}" target="_blank" style="background-color: #6c757d; color: white; padding: 4px 8px; border-radius: 4px; text-decoration: none; font-size: 11px; font-weight: bold;">🧾 Factura</a>', obj.factura_proveedor.url)
        return "-"
    ver_factura.short_description = "Factura"

    def documentacion_ok(self, obj):
        if obj.documentacion_completa: return mark_safe('<span style="color: #28a745; font-weight: bold;">✔ Completa</span>')
        return mark_safe('<span style="color: #dc3545; font-weight: bold;">❌ Faltante</span>')
    documentacion_ok.short_description = "Documentación"


# ==========================================
# 🚚 RECUPERADO: TRASPASOS INTERCOMPANY Y ESCÁNER KARDEX
# ==========================================

@admin.register(TraspasoIntercompany)
class TraspasoIntercompanyAdmin(admin.ModelAdmin):
    list_display = ('folio_factura', 'almacen_origen', 'almacen_destino', 'medicamento', 'lote', 'custom_cantidad', 'mostrar_importe', 'estatus')
    search_fields = ('folio_factura', 'medicamento__clave_sector', 'lote')
    list_filter = ('estatus', 'almacen_origen', 'almacen_destino', 'fecha_operacion')
    autocomplete_fields = ['medicamento']
    
    def custom_cantidad(self, obj):
        return f"{obj.cantidad:,}" if obj.cantidad else "0"
    custom_cantidad.short_description = "Cantidad"

    def get_readonly_fields(self, request, obj=None):
        if obj and obj.procesado: return ('almacen_origen', 'almacen_destino', 'medicamento', 'lote', 'cantidad', 'precio_unitario', 'folio_factura', 'estatus')
        return ()

    def mostrar_importe(self, obj):
        total = obj.cantidad * obj.precio_unitario
        return format_html('<b>${}</b>', f"{total:,.2f}")
    mostrar_importe.short_description = "Valor Fiscal Total"


@admin.register(EscanerKardex)
class EscanerKardexAdmin(admin.ModelAdmin):
    def changelist_view(self, request, extra_context=None):
        return HttpResponseRedirect(reverse('buscar_kardex'))
    
from .models import HistorialPrecio

from django.contrib import admin
from django.utils.html import format_html
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from .models import HistorialPrecio, CatalogoMedicamento

# ==========================================
# 📊 PANEL: HISTORIAL DE PRECIOS
# ==========================================

# 1. TRADUCTOR
class HistorialPrecioResource(resources.ModelResource):
    medicamento = fields.Field(
        column_name='clave',
        attribute='medicamento',
        widget=MedicamentoSeguroWidget(CatalogoMedicamento, 'clave_sector')
    )

    class Meta:
        model = HistorialPrecio
        import_id_fields = ('clave', 'procedimiento')

# 2. PANEL VISUAL (Aquí estaba el error del letrero)
@admin.register(HistorialPrecio)
class HistorialPrecioAdmin(ImportExportModelAdmin):
    resource_class = HistorialPrecioResource
    
    list_per_page = 50
    list_display = ('clave', 'descripcion_corta', 'institucion', 'procedimiento', 'proveedor', 'mostrar_precio', 'cantidad', 'mostrar_valor', 'fecha_captura')
    search_fields = ('clave', 'descripcion', 'procedimiento', 'institucion', 'proveedor')
    list_filter = ('institucion', 'tipo_procedimiento', 'proveedor', 'fecha_captura')

    def descripcion_corta(self, obj):
        if obj.descripcion and len(obj.descripcion) > 35:
            return f"{obj.descripcion[:35]}..."
        return obj.descripcion
    descripcion_corta.short_description = "Descripción"

    def mostrar_precio(self, obj):
        return format_html('<b style="color: #28a745;">${:,.2f}</b>', obj.precio)
    mostrar_precio.short_description = "Precio Unitario"
    mostrar_precio.admin_order_field = 'precio'

    def mostrar_valor(self, obj):
        return format_html('<b>${:,.2f}</b>', obj.valor)
    mostrar_valor.short_description = "Valor Total"
    mostrar_valor.admin_order_field = 'valor'

    def has_import_permission(self, request):
        return True
        
    def has_add_permission(self, request):
        return False

    def get_readonly_fields(self, request, obj=None):
        return [f.name for f in self.model._meta.fields]